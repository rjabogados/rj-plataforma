import traceback
import csv
import openpyxl
import uuid
import json
import random 
import secrets
import string
import mimetypes
import os
import re
import unicodedata
from urllib.parse import quote
from datetime import datetime, date, timedelta
from pathlib import Path
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from django.contrib.auth.models import User
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Q, Prefetch, Avg, Count, Max
from django.utils import timezone
from django.http import HttpResponse, FileResponse, Http404
from django.views.decorators.http import require_http_methods

from intranet.models import (
    Colaborador, Negocio, Area, Cargo, Encuesta, Pregunta, RespuestaEncuesta,
    MensajeInterno, EventoCalendario, Comunicado, CandidatoOnboarding,
    MaterialFormativo, MatriculaCurso,
    PreguntaEvaluacion, RespuestaColaborador, OpcionRespuesta, CategoriaModuloLMS, OpcionPregunta,
    RutaInduccion, RutaInduccionModulo, Notificacion
)

from .utils import solo_directivos, solo_calidad, generar_username_unico, filtrar_colaboradores, filtros_personal_disponibles, perfil_coincide_segmentacion

from intranet.models.lms import EvaluacionCurso, CursoInduccion, LeccionCurso, ProgresoLeccion

MAX_INTERNO_ADJUNTO_SIZE = 15 * 1024 * 1024
ADJUNTOS_COMUNICACION_PERMITIDOS = {
    '.pdf', '.doc', '.docx', '.xls', '.xlsx', '.ppt', '.pptx', '.txt',
    '.jpg', '.jpeg', '.png', '.gif', '.webp', '.mp4', '.mov', '.avi', '.webm'
}
MAX_PORTADA_CURSO_SIZE = 2 * 1024 * 1024
PORTADA_CURSO_EXT_PERMITIDAS = {'.jpg', '.jpeg', '.png', '.webp'}
ROLE_AREA_MAP = {
    'ASESOR': 'Operaciones',
    'BACKOFFICE': 'Operaciones',
    'CALIDAD': 'Calidad',
    'SUPERVISOR': 'Calidad',
    'SISTEMAS': 'Sistemas',
    'ADMINISTRATIVO': 'Administracion',
    'RRHH': 'Recursos Humanos',
    'GERENCIA': 'Gerencia',
}


def ruta_compatible_con_colaborador(ruta, colaborador):
    if ruta.rol_objetivo and ruta.rol_objetivo != colaborador.rol:
        return False
    if ruta.area_objetivo_id and ruta.area_objetivo_id != colaborador.area_id:
        return False
    if ruta.cartera_objetivo_id and ruta.cartera_objetivo_id != colaborador.negocio_id:
        return False
    if ruta.subcartera_objetivo and (ruta.subcartera_objetivo.lower() != (colaborador.subcartera or '').lower()):
        return False
    return True


def _usuarios_notificacion_encuestas():
    return User.objects.filter(
        perfil__rol__in=['RRHH', 'ADMINISTRATIVO', 'GERENCIA']
    ).distinct()


def _notificar_respuesta_encuesta(encuesta, perfil):
    if not perfil:
        return

    if encuesta.es_anonima:
        detalle = f'Se registrÃ³ una respuesta anÃ³nima en "{encuesta.titulo}".'
    else:
        detalle = f'{perfil.user.get_full_name() or perfil.user.username} respondiÃ³ "{encuesta.titulo}".'

    for usuario in _usuarios_notificacion_encuestas():
        Notificacion.objects.create(
            usuario=usuario,
            tipo='ENCUESTA',
            titulo=f'Nueva respuesta de encuesta: {encuesta.titulo}',
            detalle=detalle,
            url_destino='/encuestas-control/',
        )


def generar_codigo_certificado(matricula):
    if matricula.certificado_codigo:
        return matricula.certificado_codigo
    return f"RJ-{matricula.id}-{uuid.uuid4().hex[:8].upper()}"


def curso_prerequisito_cumplido(colaborador, curso):
    if not curso.prerequisito_curso_id:
        return True
    return MatriculaCurso.objects.filter(
        colaborador=colaborador,
        curso_id=curso.prerequisito_curso_id,
        estado='COMPLETADO'
    ).exists()


def usuario_es_directivo(user):
    if user.is_superuser:
        return True
    perfil = getattr(user, 'perfil', None)
    return bool(perfil and perfil.rol in ['ADMINISTRATIVO', 'RRHH', 'GERENCIA'])


def _destino_panel_por_tipo(tipo_curso):
    return 'onboarding_admin' if tipo_curso == 'INDUCCION' else 'gestor_lms'


def _leccion_tiene_contenido_minimo(url_video, archivo_pdf):
    return bool((url_video or '').strip() or archivo_pdf)


def build_storage_response(field_file):
    if not field_file or not field_file.name:
        raise Http404("Archivo no disponible")

    storage = field_file.storage
    if not storage.exists(field_file.name):
        raise Http404("Archivo no disponible")

    filename = field_file.name.split('/')[-1]
    content_type, _ = mimetypes.guess_type(filename)
    response = FileResponse(storage.open(field_file.name, 'rb'), content_type=content_type or 'application/octet-stream')
    response['Content-Disposition'] = f"inline; filename*=UTF-8''{quote(filename)}"
    return response


def adjunto_comunicacion_valido(adjunto):
    if not adjunto:
        return True
    extension = os.path.splitext(adjunto.name)[1].lower()
    return extension in ADJUNTOS_COMUNICACION_PERMITIDOS and adjunto.size <= MAX_INTERNO_ADJUNTO_SIZE


def portada_curso_valida(portada):
    if not portada:
        return True, ''
    extension = Path(portada.name).suffix.lower()
    if extension not in PORTADA_CURSO_EXT_PERMITIDAS:
        return False, 'La miniatura debe ser JPG, PNG o WEBP.'
    if portada.size > MAX_PORTADA_CURSO_SIZE:
        return False, 'La miniatura supera el mÃ¡ximo permitido (2 MB).'
    return True, ''


def subcarteras_catalogo():
    fijas = ['Delfos', 'Impaga', 'Particulares/Pyme', 'Mora Temprana', 'Especiales', 'Castigada', 'Judicial']
    valores = Colaborador.objects.exclude(subcartera__isnull=True).exclude(subcartera__exact='').values_list('subcartera', flat=True)
    unicos = {v.casefold(): v for v in fijas}
    for valor in valores:
        limpio = ' '.join(str(valor).split())
        if not limpio:
            continue
        llave = limpio.casefold()
        if llave not in unicos:
            unicos[llave] = limpio
    return [unicos[k] for k in sorted(unicos.keys())]


def _normalizar_texto(valor):
    texto = str(valor or '').strip()
    texto = ' '.join(texto.split())
    if not texto:
        return ''
    base = unicodedata.normalize('NFKD', texto)
    base = ''.join(c for c in base if unicodedata.category(c) != 'Mn')
    return base.lower()


def _detectar_rol(rol_raw):
    rol_text = _normalizar_texto(rol_raw)
    if not rol_text:
        return 'ASESOR'

    aliases = [
        ('RRHH', ['rrhh', 'recursos humanos', 'talento humano']),
        ('GERENCIA', ['gerencia', 'gerente', 'direccion']),
        ('SISTEMAS', ['sistemas', 'it', 'tecnologia', 'soporte tecnico']),
        ('SUPERVISOR', ['supervisor', 'lider', 'coordinador']),
        ('CALIDAD', ['calidad', 'qa', 'monitor']),
        ('BACKOFFICE', ['backoffice', 'back office', 'operaciones back']),
        ('ADMINISTRATIVO', ['administrativo', 'administracion', 'adm']),
        ('ASESOR', ['asesor', 'ejecutivo', 'gestor', 'cobrador', 'teleoperador']),
    ]
    for rol, palabras in aliases:
        if any(p in rol_text for p in palabras):
            return rol

    for rol_key, rol_label in Colaborador.ROLES:
        if rol_text == _normalizar_texto(rol_key) or rol_text == _normalizar_texto(rol_label):
            return rol_key
    return 'ASESOR'


def _detectar_tipo_horario(valor):
    raw = _normalizar_texto(valor)
    if not raw:
        return 'T1'
    if raw in ['t1', 'turno manana', 'manana']:
        return 'T1'
    if raw in ['t2', 'turno tarde', 'tarde']:
        return 'T2'
    if raw in ['tc', 'turno completo', 'completo', 'full time']:
        return 'TC'
    if raw in ['pt', 'part time', 'medio tiempo']:
        return 'PT'
    return 'T1'


def _partes_texto_estructura(*valores):
    texto = ' | '.join([str(v or '').strip() for v in valores if str(v or '').strip()])
    if not texto:
        return []
    partes = [p.strip() for p in re.split(r'\||/|;|,|>|\\|-', texto) if p and p.strip()]
    return partes


def _resolver_estructura_desde_texto(partes, rol_key):
    negocios = list(Negocio.objects.all().order_by('nombre'))
    areas = list(Area.objects.filter(activa=True).order_by('nombre'))
    cargos = list(Cargo.objects.filter(activa=True).select_related('area').order_by('nombre'))

    negocio_obj = None
    area_obj = None
    cargo_obj = None
    subcartera = None

    for parte in partes:
        n_parte = _normalizar_texto(parte)
        if not n_parte:
            continue
        if not negocio_obj:
            negocio_obj = next((n for n in negocios if _normalizar_texto(n.nombre) in n_parte or n_parte in _normalizar_texto(n.nombre)), None)
            if negocio_obj:
                continue
        if not area_obj:
            area_obj = next((a for a in areas if _normalizar_texto(a.nombre) in n_parte or n_parte in _normalizar_texto(a.nombre)), None)
            if area_obj:
                continue
        if not cargo_obj:
            cargo_obj = next((c for c in cargos if _normalizar_texto(c.nombre) in n_parte or n_parte in _normalizar_texto(c.nombre)), None)
            if cargo_obj:
                continue
        if not subcartera and len(parte) >= 3:
            subcartera = ' '.join(parte.split())

    if not area_obj:
        area_nombre = ROLE_AREA_MAP.get(rol_key, 'General')
        area_obj = Area.objects.filter(nombre__iexact=area_nombre).first()
        if not area_obj:
            area_obj = Area.objects.create(nombre=area_nombre, activa=True)

    if cargo_obj and cargo_obj.area and not area_obj:
        area_obj = cargo_obj.area

    if not cargo_obj:
        cargo_nombre = dict(Colaborador.ROLES).get(rol_key, 'Colaborador')
        cargo_obj = Cargo.objects.filter(area=area_obj, nombre__iexact=cargo_nombre).first()
        if not cargo_obj:
            cargo_obj = Cargo.objects.create(area=area_obj, nombre=cargo_nombre, activa=True)

    return negocio_obj, area_obj, cargo_obj, subcartera


def _nombre_apellido_desde_texto(nombre_raw, apellido_raw, separados):
    if separados:
        return (str(nombre_raw or '').strip().title(), str(apellido_raw or '').strip().title())

    combinado = ' '.join(str(nombre_raw or '').split()).strip()
    if ',' in combinado:
        apellidos, nombres = [p.strip() for p in combinado.split(',', 1)]
        return (nombres.title(), apellidos.title())

    piezas = combinado.split()
    if len(piezas) >= 3:
        return (' '.join(piezas[2:]).title(), ' '.join(piezas[:2]).title())
    if len(piezas) == 2:
        return (piezas[1].title(), piezas[0].title())
    return (combinado.title(), '')


def _get_excel_value(row, idx):
    if idx is None or idx < 0 or idx >= len(row):
        return ''
    val = row[idx]
    return str(val).strip() if val is not None else ''


def _parse_index(value):
    if value in [None, '']:
        return None
    return int(value)


def es_participante_mensaje(mensaje, perfil):
    return mensaje.remitente_id == perfil.id or mensaje.destinatario_id == perfil.id


def crear_comunicado_desde_request(request):
    titulo = request.POST.get('titulo', '').strip()
    mensaje = request.POST.get('mensaje', '').strip()
    adjunto = request.FILES.get('adjunto')

    if not titulo or not mensaje:
        messages.error(request, 'El tÃ­tulo y el mensaje son obligatorios.')
        return False
    if not adjunto_comunicacion_valido(adjunto):
        messages.error(request, 'El adjunto no es vÃ¡lido. Usa archivos permitidos de hasta 15 MB.')
        return False

    Comunicado.objects.create(titulo=titulo[:200], mensaje=mensaje, adjunto=adjunto, activo=True)
    messages.success(request, 'Comunicado publicado correctamente.')
    return True

# ==========================================
# DIRECTORIO DE PERSONAL E IMPORTACIÃ“N EXCEL
# ==========================================
@login_required(login_url='login')
@solo_directivos
def colaboradores(request):
    if not Area.objects.exists():
        from intranet.utils.poblar import poblar_taxonomia_completa
        poblar_taxonomia_completa()
        
    perfil_actual = getattr(request.user, 'perfil', None)

    if request.method == 'POST':
        nombres = request.POST.get('nombres')
        apellidos = request.POST.get('apellidos')
        dni_val = (request.POST.get('dni') or '').strip() or None
        correo_val = request.POST.get('correo').strip().lower() or None
        rol_val = request.POST.get('rol')
        negocio_id = request.POST.get('negocio')
        area_id = request.POST.get('area')
        cargo_id = request.POST.get('cargo')
        subcartera = request.POST.get('subcartera', '').strip() or None
        tipo_horario = request.POST.get('tipo_horario')
        
        sede_val = request.POST.get('sede') or 'LIMA1'
        username_custom = request.POST.get('username', '').strip()
        password_custom = request.POST.get('password', '').strip()

        username_final = username_custom if username_custom else generar_username_unico(nombres, apellidos, dni_val)
        password_final = password_custom if password_custom else generar_contrasena_segura()

        negocio_instancia = Negocio.objects.get(id=negocio_id) if negocio_id else None
        area_instancia = Area.objects.filter(id=area_id).first() if area_id else None
        cargo_instancia = Cargo.objects.select_related('area').filter(id=cargo_id).first() if cargo_id else None
        if cargo_instancia and not area_instancia and cargo_instancia.area:
            area_instancia = cargo_instancia.area
        f_ingreso = request.POST.get('fecha_ingreso')
        fecha_formal = datetime.strptime(f_ingreso, '%Y-%m-%d').date() if f_ingreso else date.today()

        if not User.objects.filter(username=username_final).exists():
            nuevo_user = User.objects.create_user(
                username=username_final, email=correo_val if correo_val else "",
                password=password_final, first_name=nombres, last_name=apellidos
            )
            Colaborador.objects.create(
                user=nuevo_user, dni=dni_val, rol=rol_val, sede=sede_val, negocio=negocio_instancia, 
                area=area_instancia, cargo=cargo_instancia, subcartera=subcartera,
                tipo_horario=tipo_horario, hora_ingreso=request.POST.get('hora_ingreso') or None, 
                hora_salida=request.POST.get('hora_salida') or None, fecha_ingreso=fecha_formal
            )
            return redirect('colaboradores')

    lista_colaboradores = filtrar_colaboradores(
        Colaborador.objects.select_related('user', 'negocio', 'area', 'cargo'),
        request.GET,
        perfil_actual,
    ).order_by('area__nombre', 'cargo__nombre', 'user__last_name', 'user__first_name')

    return render(request, 'intranet/rrhh/colaboradores.html', {
        'colaboradores': lista_colaboradores,
        'negocios': Negocio.objects.all().order_by('nombre'),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).select_related('area').order_by('area__nombre', 'nombre'),
        'roles': Colaborador.ROLES,
        'sedes': Colaborador.SEDES,
        'tipos_horario': Colaborador.TIPO_HORARIO,
        'subcarteras': subcarteras_catalogo(),
        'filtros_disponibles': filtros_personal_disponibles(perfil_actual),
    })

def generar_contrasena_segura():
    """Genera una contraseña aleatoria de 12 caracteres"""
    alphabet = string.ascii_letters + string.digits + string.punctuation
    # Evitar caracteres problemÃ¡ticos
    alphabet = alphabet.replace("'", "").replace('"', "").replace('\\', "")
    contrasena = ''.join(secrets.choice(alphabet) for i in range(12))
    return contrasena


@login_required(login_url='login')
@solo_directivos
def editar_colaborador(request, pk):
    colab = get_object_or_404(Colaborador, pk=pk)
    
    if request.method == 'POST':
        colab.user.first_name = request.POST.get('nombres')
        colab.user.last_name = request.POST.get('apellidos')
        colab.user.email = request.POST.get('correo').strip().lower() or ""
        
        nuevo_username = request.POST.get('username', '').strip()
        nueva_password = request.POST.get('password', '').strip()

        if nuevo_username and not User.objects.filter(username=nuevo_username).exclude(pk=colab.user.pk).exists():
            colab.user.username = nuevo_username
            
        if nueva_password:
            colab.user.set_password(nueva_password)

        colab.user.save()
        colab.dni = (request.POST.get('dni') or '').strip() or None
        colab.rol = request.POST.get('rol')
        colab.tipo_horario = request.POST.get('tipo_horario')
        colab.hora_ingreso = request.POST.get('hora_ingreso') or None
        colab.hora_salida = request.POST.get('hora_salida') or None
        if request.POST.get('fecha_ingreso'):
            colab.fecha_ingreso = datetime.strptime(request.POST.get('fecha_ingreso'), '%Y-%m-%d').date()
        
        colab.sede = request.POST.get('sede')
        negocio_id = request.POST.get('negocio')
        colab.negocio = Negocio.objects.get(id=negocio_id) if negocio_id else None
        area_id = request.POST.get('area')
        cargo_id = request.POST.get('cargo')
        colab.area = Area.objects.filter(id=area_id).first() if area_id else None
        colab.cargo = Cargo.objects.select_related('area').filter(id=cargo_id).first() if cargo_id else None
        if colab.cargo and not colab.area and colab.cargo.area:
            colab.area = colab.cargo.area
        colab.subcartera = request.POST.get('subcartera', '').strip() or None
        colab.save()

        onboarding_activo = request.POST.get('switch_onboarding') == 'on'
        
        if onboarding_activo:
            if colab.dni:
                CandidatoOnboarding.objects.get_or_create(
                    colaborador=colab, dni=colab.dni,
                    defaults={
                        'nombres': colab.user.first_name,
                        'apellidos': colab.user.last_name,
                        'estado': 'EN_PROCESO'
                    }
                )
            else:
                messages.warning(request, 'No se activÃ³ Onboarding porque el colaborador no tiene DNI/documento registrado.')
        else:
            CandidatoOnboarding.objects.filter(colaborador=colab).delete()

        return redirect('colaboradores')
        
    tiene_onboarding = CandidatoOnboarding.objects.filter(colaborador=colab).exists()
    return render(request, 'intranet/rrhh/editar_colaborador.html', {
        'colab': colab,
        'negocios': Negocio.objects.all().order_by('nombre'),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).select_related('area').order_by('area__nombre', 'nombre'),
        'roles': Colaborador.ROLES,
        'sedes': Colaborador.SEDES,
        'tipos_horario': Colaborador.TIPO_HORARIO,
        'subcarteras': subcarteras_catalogo(),
        'tiene_onboarding': tiene_onboarding
    })

@login_required(login_url='login')
@solo_directivos
def eliminar_colaborador(request, pk):
    colab = get_object_or_404(Colaborador, pk=pk)
    user_vinculado = colab.user
    colab.delete()
    if user_vinculado: user_vinculado.delete()
    return redirect('colaboradores')

@login_required(login_url='login')
@solo_directivos
def mapear_excel(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        nombre_tmp = default_storage.save(f'tmp/{request.user.id}_import.xlsx', ContentFile(excel_file.read()))
        wb = openpyxl.load_workbook(default_storage.open(nombre_tmp))
        cabeceras_excel = [str(celda.value).strip() for celda in wb.active[1] if celda.value is not None]
        request.session['ruta_excel_tmp'] = nombre_tmp
        return render(request, 'intranet/documentos/mapear_excel.html', {'cabeceras': cabeceras_excel})
    return redirect('colaboradores')


@login_required(login_url='login')
@solo_directivos
def procesar_mapeo_personal(request):
    if request.method != 'POST':
        return redirect('colaboradores')

    ruta_archivo = request.session.get('ruta_excel_tmp')
    if not ruta_archivo or not default_storage.exists(ruta_archivo):
        messages.error(request, 'El archivo de personal expirÃ³. Sube el Excel nuevamente.')
        return redirect('colaboradores')

    try:
        idx_dni = _parse_index(request.POST.get('prop_dni'))
        idx_correo = _parse_index(request.POST.get('prop_correo'))
        idx_nombres = _parse_index(request.POST.get('prop_nombres'))
        idx_apellidos = _parse_index(request.POST.get('prop_apellidos'))
        idx_rol = _parse_index(request.POST.get('prop_rol'))
        idx_tipo_horario = _parse_index(request.POST.get('prop_tipo_horario'))
        idx_cartera = _parse_index(request.POST.get('prop_cartera'))
        idx_subcartera = _parse_index(request.POST.get('prop_subcartera'))
        idx_estructura = _parse_index(request.POST.get('prop_estructura'))
        separar_nombres = request.POST.get('separar_nombres') == '1'

        archivo_excel = default_storage.open(ruta_archivo)
        wb = openpyxl.load_workbook(archivo_excel, data_only=True, read_only=True)
        ws = wb.active

        creados = 0
        actualizados = 0
        omitidos = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            dni_raw = _get_excel_value(row, idx_dni)
            dni = ''.join(ch for ch in dni_raw if ch.isalnum()) or None
            if not dni:
                omitidos += 1
                continue

            correo = _get_excel_value(row, idx_correo).lower()
            nombre_raw = _get_excel_value(row, idx_nombres)
            apellido_raw = _get_excel_value(row, idx_apellidos)
            nombres, apellidos = _nombre_apellido_desde_texto(nombre_raw, apellido_raw, separar_nombres)

            rol_raw = _get_excel_value(row, idx_rol)
            rol_key = _detectar_rol(rol_raw)
            tipo_horario = _detectar_tipo_horario(_get_excel_value(row, idx_tipo_horario))

            cartera_raw = _get_excel_value(row, idx_cartera)
            subcartera_raw = _get_excel_value(row, idx_subcartera)
            estructura_libre = _get_excel_value(row, idx_estructura)
            partes = _partes_texto_estructura(estructura_libre, rol_raw, cartera_raw, subcartera_raw)
            negocio_obj, area_obj, cargo_obj, subcartera_inferida = _resolver_estructura_desde_texto(partes, rol_key)

            subcartera_final = ' '.join((subcartera_raw or subcartera_inferida or '').split()) or None

            colab_existente = Colaborador.objects.select_related('user').filter(dni=dni).first()
            if colab_existente:
                user = colab_existente.user
                user.first_name = nombres or user.first_name
                user.last_name = apellidos or user.last_name
                if correo:
                    user.email = correo
                user.save(update_fields=['first_name', 'last_name', 'email'])
            else:
                username_base = generar_username_unico(nombres or 'usuario', apellidos or 'rj', dni)
                user = User.objects.create_user(
                    username=username_base,
                    email=correo,
                    password=generar_contrasena_segura(),
                    first_name=nombres,
                    last_name=apellidos,
                )

            colab, creado_colab = Colaborador.objects.update_or_create(
                dni=dni,
                defaults={
                    'user': user,
                    'rol': rol_key,
                    'negocio': negocio_obj,
                    'area': area_obj,
                    'cargo': cargo_obj,
                    'subcartera': subcartera_final,
                    'tipo_horario': tipo_horario,
                    'fecha_ingreso': date.today(),
                }
            )

            if creado_colab:
                creados += 1
            else:
                actualizados += 1

        wb.close()
        archivo_excel.close()
        default_storage.delete(ruta_archivo)
        request.session.pop('ruta_excel_tmp', None)

        messages.success(
            request,
            f'Base personal procesada. Creados: {creados}, actualizados: {actualizados}, omitidos: {omitidos}.'
        )
    except Exception:
        messages.error(request, 'OcurriÃ³ un error al procesar la base de personal. Verifica el mapeo de columnas.')

    return redirect('colaboradores')

@login_required(login_url='login')
def procesar_mapeo_balotario(request):
    if request.method == 'POST':
        try:
            destino = request.session.get('balotario_return_to', 'gestor_lms')
            ruta_archivo = request.session.get('ruta_excel_balotario')
            if not ruta_archivo or not default_storage.exists(ruta_archivo):
                messages.error(request, "El archivo expirÃ³. Vuelve a subirlo.")
                return redirect(destino)

            idx_pregunta = int(request.POST.get('prop_pregunta', -1))
            idx_correcta = int(request.POST.get('prop_correcta', -1))
            idx_alt1 = int(request.POST.get('prop_alt1', -1))
            idx_alt2 = int(request.POST.get('prop_alt2', -1))
            idx_alt3 = int(request.POST.get('prop_alt3', -1))
            idx_alt4 = int(request.POST.get('prop_alt4', -1))

            archivo_excel = default_storage.open(ruta_archivo)
            wb = openpyxl.load_workbook(archivo_excel, data_only=True, read_only=True)
            
            preguntas_temporales = []
            for i, fila in enumerate(wb.active.iter_rows(min_row=2, values_only=True)):
                def get_val(idx):
                    if idx != -1 and idx < len(fila):
                        val = fila[idx]
                        return str(val).strip() if val is not None else ""
                    return ""

                enunciado = get_val(idx_pregunta)
                if enunciado: 
                    preguntas_temporales.append({
                        'id_temp': int(i),
                        'enunciado': enunciado,
                        'correcta': get_val(idx_correcta),
                        'alt1': get_val(idx_alt1), 
                        'alt2': get_val(idx_alt2), 
                        'alt3': get_val(idx_alt3), 
                        'alt4': get_val(idx_alt4)
                    })

            wb.close()
            archivo_excel.close()
            try:
                default_storage.delete(ruta_archivo)
            except Exception: pass 
            
            if 'ruta_excel_balotario' in request.session:
                del request.session['ruta_excel_balotario']

            request.session['balotario_temporal'] = preguntas_temporales
            request.session.save()
            return redirect('previsualizar_balotario')
        except Exception as e:
            error_texto = traceback.format_exc()
            return HttpResponse(f"<div style='padding:20px; font-family: monospace; background:#ffe6e6; color:red; border:2px solid red;'><h2>Â¡EL ERROR FUE ATRAPADO!</h2><pre>{error_texto}</pre></div>", status=200)
    return redirect(request.session.get('balotario_return_to', 'gestor_lms'))


# ==========================================
# ONBOARDING CORPORATIVO (INDUCCIÃ“N)
# ==========================================
@login_required(login_url='login')
def induccion(request): return redirect('mi_induccion')

@login_required(login_url='login')
@solo_calidad
def induccion_admin(request): return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def gestionar_onboarding(request): return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def onboarding_admin(request):
    from intranet.models.rrhh_core import Negocio, Area, Cargo, Colaborador
    from intranet.models.lms import CursoInduccion, EvaluacionCurso, LeccionCurso, CandidatoOnboarding
    
    if request.method == 'POST':
        if 'crear_ruta_induccion' in request.POST:
            nombre_ruta = (request.POST.get('nombre_ruta') or '').strip()
            if not nombre_ruta:
                messages.error(request, 'La ruta de inducciÃ³n necesita un nombre.')
                return redirect('onboarding_admin')

            RutaInduccion.objects.create(
                nombre=nombre_ruta,
                descripcion=(request.POST.get('descripcion_ruta') or '').strip(),
                rol_objetivo=request.POST.get('rol_objetivo') or None,
                cartera_objetivo=Negocio.objects.filter(id=request.POST.get('cartera_objetivo')).first() if request.POST.get('cartera_objetivo') else None,
                subcartera_objetivo=(request.POST.get('subcartera_objetivo') or '').strip() or None,
                area_objetivo=Area.objects.filter(id=request.POST.get('area_objetivo')).first() if request.POST.get('area_objetivo') else None,
                activa=True,
            )
            messages.success(request, 'Ruta de inducciÃ³n creada correctamente.')
            return redirect('onboarding_admin')

        if 'agregar_modulo_ruta' in request.POST:
            ruta = get_object_or_404(RutaInduccion, id=request.POST.get('ruta_id'))
            modulo = get_object_or_404(CursoInduccion, id=request.POST.get('modulo_id'), tipo='INDUCCION')
            orden = int(request.POST.get('orden') or 1)
            item, created = RutaInduccionModulo.objects.get_or_create(
                ruta=ruta,
                modulo=modulo,
                defaults={'orden': orden},
            )
            if not created:
                item.orden = orden
                item.save(update_fields=['orden'])
            prereq_id = request.POST.get('prerequisito_item')
            item.prerequisito = RutaInduccionModulo.objects.filter(id=prereq_id, ruta=ruta).first() if prereq_id else None
            item.save(update_fields=['prerequisito'])
            messages.success(request, 'MÃ³dulo agregado/actualizado en la ruta de inducciÃ³n.')
            return redirect('onboarding_admin')

        # 1. CREAR MÃ“DULO DE INDUCCIÃ“N
        if 'crear_modulo' in request.POST:
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            
            # --- Smart Targeting (SegmentaciÃ³n) ---
            publico_general = request.POST.get('publico_general') == 'on'
            rol_permitido = request.POST.get('rol_permitido') or None
            area_permitida_id = request.POST.get('area_permitida')
            cargo_permitido_id = request.POST.get('cargo_permitido')
            cartera_id = request.POST.get('cartera_vinculada')
            subcartera = request.POST.get('subcartera_vinculada') or None
            prerequisito_id = request.POST.get('prerequisito_curso')
            version_val = int(request.POST.get('version') or 1)
            
            cartera_obj = Negocio.objects.filter(id=cartera_id).first() if cartera_id else None
            area_obj = Area.objects.filter(id=area_permitida_id).first() if area_permitida_id else None
            cargo_obj = Cargo.objects.select_related('area').filter(id=cargo_permitido_id).first() if cargo_permitido_id else None
            prerequisito_obj = CursoInduccion.objects.filter(id=prerequisito_id, tipo='INDUCCION').first() if prerequisito_id else None
            if cargo_obj and not area_obj and cargo_obj.area:
                area_obj = cargo_obj.area

            CursoInduccion.objects.create(
                titulo=titulo, descripcion=descripcion, tipo='INDUCCION',
                publico_general=publico_general, rol_permitido=rol_permitido,
                area_permitida=area_obj, cargo_permitido=cargo_obj,
                cartera_vinculada=cartera_obj, subcartera_vinculada=subcartera,
                prerequisito_curso=prerequisito_obj,
                version=version_val,
                estado_publicacion=request.POST.get('estado_publicacion') or 'PUBLICADO',
            )
            messages.success(request, f"MÃ³dulo de InducciÃ³n '{titulo}' creado exitosamente.")
            
        # 2. AGREGAR LECCIÃ“N (SOPORTE CANVA / PPT)
        elif 'crear_leccion' in request.POST:
            curso_id = request.POST.get('curso_id')
            curso = get_object_or_404(CursoInduccion, id=curso_id, tipo='INDUCCION')
            
            LeccionCurso.objects.create(
                curso=curso,
                titulo=request.POST.get('titulo'),
                descripcion=request.POST.get('descripcion'),
                url_video=request.POST.get('url_video'),
                url_presentacion_canva=request.POST.get('url_presentacion_canva'), # <--- Link de Canva
                url_simulador=request.POST.get('url_simulador') or None,
                paquete_scorm_url=request.POST.get('paquete_scorm_url') or None,
                archivo_pdf=request.FILES.get('archivo_pdf'),
                orden=request.POST.get('orden', 1)
            )
            messages.success(request, "Â¡Material interactivo agregado correctamente a la inducciÃ³n!")

        # 3. CREAR PRUEBA RÃPIDA (EXAMEN)
        elif 'crear_evaluacion' in request.POST:
            curso_id = request.POST.get('curso_id')
            curso = get_object_or_404(CursoInduccion, id=curso_id, tipo='INDUCCION')
            
            if hasattr(curso, 'evaluacion'):
                messages.error(request, "Este mÃ³dulo ya tiene una prueba configurada.")
            else:
                EvaluacionCurso.objects.create(
                    curso=curso,
                    titulo=request.POST.get('titulo', 'Prueba RÃ¡pida de InducciÃ³n'),
                    puntaje_maximo=request.POST.get('puntaje_maximo', 20),
                    puntaje_aprobatorio=request.POST.get('puntaje_aprobatorio', 14),
                    preguntas_a_mostrar=request.POST.get('preguntas_a_mostrar', 5),
                    orden_aleatorio=request.POST.get('orden_aleatorio') == 'on',
                    tiempo_limite_minutos=request.POST.get('tiempo_limite_minutos', 10),
                    puntos_premio=10,
                    intentos_maximos=int(request.POST.get('intentos_maximos') or 0),
                )
                messages.success(request, "Â¡Prueba rÃ¡pida creada! Ya puedes subir tu balotario en formato Excel.")
                
        # 4. REGISTRAR CANDIDATO NUEVO
        elif 'registrar_candidato' in request.POST:
            nombres_val = request.POST.get('nombres')
            apellidos_val = request.POST.get('apellidos')
            dni_val = request.POST.get('dni')
            telefono_val = request.POST.get('telefono', '')
            puesto_val = request.POST.get('puesto_esperado', 'ASESOR')
            negocio_id = request.POST.get('campaña_destino')
            
            if CandidatoOnboarding.objects.filter(dni=dni_val).exists():
                messages.error(request, f"El DNI {dni_val} ya se encuentra registrado.")
            else:
                CandidatoOnboarding.objects.create(nombres=nombres_val, apellidos=apellidos_val, dni=dni_val, telefono=telefono_val, puesto_esperado=puesto_val, campaña_destino_id=negocio_id if negocio_id else None)
                messages.success(request, "Postulante registrado correctamente.")
        return redirect('onboarding_admin')

    # --- DATOS PARA EL DASHBOARD DE RRHH ---
    onboardings_activos = CandidatoOnboarding.objects.all().select_related('colaborador__user', 'campaña_destino')
    lista_candidatos_progreso = []
    for item in onboardings_activos:
        if item.colaborador:
            matriculas = MatriculaCurso.objects.filter(colaborador=item.colaborador)
            total_cursos = matriculas.count()
            completados = matriculas.filter(estado='COMPLETADO').count()
            porcentaje = int((completados / total_cursos) * 100) if total_cursos > 0 else 0
            ratio = f"{completados}/{total_cursos}"
        else:
            porcentaje = item.porcentaje_expediente()
            ratio = "Expediente"
        lista_candidatos_progreso.append({'onboarding': item, 'porcentaje': porcentaje, 'ratio': ratio})

    # Extraemos SOLO los de inducciÃ³n (y sus clases/exÃ¡menes precargados)
    cursos_biblioteca = CursoInduccion.objects.filter(activo=True, tipo='INDUCCION').select_related('evaluacion').prefetch_related('lecciones').order_by('-fecha_creacion')
    rutas_induccion = RutaInduccion.objects.filter(activa=True).prefetch_related('items__modulo', 'items__prerequisito').order_by('nombre')
    negocios = Negocio.objects.all()
    areas = Area.objects.filter(activa=True).order_by('nombre')
    cargos = Cargo.objects.filter(activa=True).select_related('area').order_by('nombre')

    return render(request, 'intranet/rrhh/onboarding_lista.html', {
        'candidatos_progreso': lista_candidatos_progreso, 'candidatos': onboardings_activos, 
        'modulos_biblioteca': cursos_biblioteca, 'rutas_induccion': rutas_induccion,
        'negocios': negocios, 'areas': areas, 'cargos': cargos, 'roles': Colaborador.ROLES
    })

@login_required(login_url='login')
@solo_directivos
def editar_curso_induccion(request, curso_id):
    from intranet.models.lms import CursoInduccion
    curso = get_object_or_404(CursoInduccion, id=curso_id, tipo='INDUCCION')
    
    if request.method == 'POST':
        curso.titulo = request.POST.get('titulo')
        curso.descripcion = request.POST.get('descripcion')
        curso.publico_general = request.POST.get('publico_general') == 'on'
        curso.rol_permitido = request.POST.get('rol_permitido') or None
        curso.version = int(request.POST.get('version') or curso.version)
        curso.estado_publicacion = request.POST.get('estado_publicacion') or curso.estado_publicacion
        area_id = request.POST.get('area_permitida')
        cargo_id = request.POST.get('cargo_permitido')
        curso.area_permitida = Area.objects.filter(id=area_id).first() if area_id else None
        curso.cargo_permitido = Cargo.objects.select_related('area').filter(id=cargo_id).first() if cargo_id else None
        if curso.cargo_permitido and not curso.area_permitida and curso.cargo_permitido.area:
            curso.area_permitida = curso.cargo_permitido.area
        curso.subcartera_vinculada = request.POST.get('subcartera_vinculada') or None
        cartera_id = request.POST.get('cartera_vinculada')
        curso.cartera_vinculada_id = cartera_id if cartera_id else None
        prerequisito_id = request.POST.get('prerequisito_curso')
        curso.prerequisito_curso = CursoInduccion.objects.filter(id=prerequisito_id, tipo='INDUCCION').exclude(id=curso.id).first() if prerequisito_id else None
        
        curso.save()
        messages.success(request, f"Â¡MÃ³dulo '{curso.titulo}' actualizado!")
    return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def eliminar_curso_induccion(request, curso_id):
    from intranet.models.lms import CursoInduccion
    curso = get_object_or_404(CursoInduccion, id=curso_id, tipo='INDUCCION')
    titulo = curso.titulo
    curso.delete()
    messages.success(request, f"El mÃ³dulo '{titulo}' ha sido eliminado.")
    return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def asignar_modulos_induccion(request, colab_id):
    colaborador = get_object_or_404(Colaborador, id=colab_id)
    cursos_disponibles = CursoInduccion.objects.filter(activo=True, tipo='INDUCCION').order_by('-fecha_creacion')
    rutas = RutaInduccion.objects.filter(activa=True).prefetch_related('items__modulo').order_by('nombre')
    
    if request.method == 'POST':
        cursos_seleccionados = request.POST.getlist('modulos_ids')
        ruta_id = request.POST.get('ruta_id')
        if ruta_id:
            ruta = RutaInduccion.objects.filter(id=ruta_id, activa=True).first()
            if ruta:
                cursos_seleccionados = [str(item.modulo_id) for item in ruta.items.order_by('orden')]
        MatriculaCurso.objects.filter(colaborador=colaborador).exclude(curso_id__in=cursos_seleccionados).exclude(estado='COMPLETADO').delete()
        for orden, c_id in enumerate(cursos_seleccionados, start=1):
            matricula, _ = MatriculaCurso.objects.get_or_create(colaborador=colaborador, curso_id=c_id)
            if not matricula.fecha_limite:
                matricula.fecha_limite = date.today() + timedelta(days=7 * orden)
                matricula.save(update_fields=['fecha_limite'])
        messages.success(request, f"Malla formativa de InducciÃ³n actualizada para {colaborador.user.first_name}.")
        return redirect('onboarding_admin')
        
    cursos_actuales = MatriculaCurso.objects.filter(colaborador=colaborador).values_list('curso_id', flat=True)
    return render(request, 'intranet/rrhh/asignar_modulos.html', {
        'colaborador': colaborador,
        'modulos_disponibles': cursos_disponibles,
        'modulos_actuales': cursos_actuales,
        'rutas': rutas,
    })

@login_required(login_url='login')
def mi_induccion(request):
    try:
        colaborador = request.user.perfil
    except:
        messages.error(request, "Tu usuario no tiene un perfil de trabajador asociado.")
        return redirect('inicio')

    cursos_disponibles = CursoInduccion.objects.filter(
        Q(publico_general=True) | 
        Q(rol_permitido=colaborador.rol) | 
        Q(area_permitida=colaborador.area) |
        Q(cargo_permitido=colaborador.cargo) |
        Q(cartera_vinculada=colaborador.negocio) |
        Q(subcartera_vinculada=colaborador.subcartera),
        activo=True,
        tipo='INDUCCION' 
    ).distinct().select_related('prerequisito_curso')

    ruta_activa = None
    for ruta in RutaInduccion.objects.filter(activa=True).prefetch_related('items__modulo__prerequisito_curso'):
        if ruta_compatible_con_colaborador(ruta, colaborador):
            ruta_activa = ruta
            break

    if ruta_activa and ruta_activa.items.exists():
        ids_ruta = [item.modulo_id for item in ruta_activa.items.order_by('orden')]
        cursos_disponibles = sorted(
            list(cursos_disponibles.filter(id__in=ids_ruta)),
            key=lambda c: ids_ruta.index(c.id)
        )
    else:
        cursos_disponibles = list(cursos_disponibles.order_by('orden_sugerido', 'id'))

    mis_modulos = []
    for curso in cursos_disponibles:
        matricula, created = MatriculaCurso.objects.get_or_create(
            colaborador=colaborador, 
            curso=curso,
            defaults={'estado': 'PENDIENTE'}
        )
        mis_modulos.append(matricula)

    bloqueados_ids = set()
    for matricula in mis_modulos:
        curso = matricula.curso
        if not curso_prerequisito_cumplido(colaborador, curso):
            bloqueados_ids.add(matricula.id)

    if request.method == 'POST' and 'marcar_completado' in request.POST:
        progreso_id = request.POST.get('progreso_id')
        matricula_actualizar = MatriculaCurso.objects.get(id=progreso_id, colaborador=colaborador)
        matricula_actualizar.estado = 'COMPLETADO'
        matricula_actualizar.fecha_finalizacion = timezone.now()
        matricula_actualizar.save()
        messages.success(request, f"Â¡MÃ³dulo '{matricula_actualizar.curso.titulo}' completado con Ã©xito!")
        return redirect('mi_induccion')

    total_modulos = len(mis_modulos)
    completados = sum(1 for m in mis_modulos if m.estado == 'COMPLETADO')
    porcentaje = int((completados / total_modulos) * 100) if total_modulos > 0 else 0

    return render(request, 'intranet/rrhh/mi_induccion.html', {
        'mis_modulos': mis_modulos,
        'total': total_modulos,
        'completados': completados,
        'porcentaje': porcentaje,
        'ruta_activa': ruta_activa,
        'bloqueados_ids': bloqueados_ids,
    })

@login_required(login_url='login')
@solo_directivos
def actualizar_expediente(request, candidato_id):
    candidato = get_object_or_404(CandidatoOnboarding, id=candidato_id)
    if request.method == 'POST':
        nombres = (request.POST.get('nombres') or '').strip()
        apellidos = (request.POST.get('apellidos') or '').strip()
        dni_input = ''.join(ch for ch in str(request.POST.get('dni') or '') if ch.isdigit())
        correo = (request.POST.get('correo') or '').strip().lower() or None
        telefono = (request.POST.get('telefono') or '').strip() or None
        puesto_esperado = request.POST.get('puesto_esperado') or candidato.puesto_esperado
        campaña_destino_id = request.POST.get('campaña_destino') or None

        if dni_input and len(dni_input) != 8:
            messages.error(request, 'El DNI debe tener exactamente 8 dÃ­gitos.')
            return redirect('onboarding_admin')

        candidato.nombres = nombres or candidato.nombres
        candidato.apellidos = apellidos or candidato.apellidos
        candidato.dni = dni_input or candidato.dni
        candidato.correo = correo
        candidato.telefono = telefono
        candidato.puesto_esperado = puesto_esperado
        candidato.campaña_destino_id = campaña_destino_id
        candidato.doc_cv = request.POST.get('doc_cv') == 'on'
        candidato.doc_dni = request.POST.get('doc_dni') == 'on'
        candidato.doc_antecedentes = request.POST.get('doc_antecedentes') == 'on'
        candidato.doc_recibo_servicios = request.POST.get('doc_recibo_servicios') == 'on'
        candidato.save()

        if candidato.colaborador:
            candidato.colaborador.user.first_name = candidato.nombres
            candidato.colaborador.user.last_name = candidato.apellidos
            if candidato.correo is not None:
                candidato.colaborador.user.email = candidato.correo
            candidato.colaborador.user.save(update_fields=['first_name', 'last_name', 'email'])
            if candidato.dni:
                candidato.colaborador.dni = candidato.dni
            candidato.colaborador.rol = candidato.puesto_esperado
            candidato.colaborador.negocio = candidato.campaña_destino
            candidato.colaborador.save(update_fields=['dni', 'rol', 'negocio'])

        messages.success(request, f"Expediente de {candidato.nombres} actualizado.")
    return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def pasar_a_planilla(request, candidato_id):
    candidato = get_object_or_404(CandidatoOnboarding, id=candidato_id)
    dni_limpio = ''.join(ch for ch in str(candidato.dni or '') if ch.isdigit())
    if len(dni_limpio) != 8:
        messages.error(request, 'El candidato no puede pasar a planilla sin un DNI vÃ¡lido de 8 dÃ­gitos.')
        return redirect('onboarding_admin')

    if candidato.porcentaje_expediente() < 100:
        messages.error(request, "Expediente incompleto. Faltan documentos.")
        return redirect('onboarding_admin')
    try:
        with transaction.atomic():
            if candidato.colaborador:
                candidato.estado = 'COMPLETADO'
                candidato.save()
            else:
                nombres_base = (candidato.nombres or 'usuario').split()
                apellidos_base = (candidato.apellidos or 'usuario').split()
                username_final = f"{nombres_base[0].lower()}.{apellidos_base[0].lower()}"
                if User.objects.filter(username=username_final).exists():
                    username_final = f"{username_final}{dni_limpio[-2:]}"
                nuevo_user = User.objects.create_user(username=username_final, email=candidato.correo or '', password=dni_limpio, first_name=candidato.nombres, last_name=candidato.apellidos)
                nuevo_colaborador = Colaborador.objects.create(
                    user=nuevo_user,
                    dni=dni_limpio,
                    rol=candidato.puesto_esperado,
                    negocio=candidato.campaña_destino,
                    fecha_ingreso=date.today(),
                )
                candidato.colaborador = nuevo_colaborador
                candidato.estado = 'COMPLETADO'
                candidato.save()
                messages.success(request, f"Â¡{candidato.nombres} ingresÃ³ a planilla!")
    except Exception:
        messages.error(request, "Error al procesar el alta. Verifique el DNI.")
    return redirect('onboarding_admin')

@login_required(login_url='login')
@solo_directivos
def organigrama_empresa(request):
    from intranet.models.rrhh_core import Area, Cargo, Negocio, Colaborador
    import json
    
    # 1. Administrativos
    areas = Area.objects.filter(activa=True).prefetch_related('cargos', 'colaboradores').order_by('nombre')
    
    admin_data = []
    for a in areas:
        cargos = a.cargos.filter(activa=True).order_by('nombre')
        cargos_list = []
        for c in cargos:
            colabs = list(c.colaboradores.values_list('user__first_name', 'user__last_name'))
            cargos_list.append({
                'nombre': c.nombre,
                'personas': [f"{fn} {ln}" for fn, ln in colabs]
            })
        
        admin_data.append({
            'nombre': a.nombre,
            'cargos': cargos_list
        })
        
    # 2. Asesores (Carteras)
    negocios = Negocio.objects.all().order_by('nombre')
    asesores_data = []
    for n in negocios:
        colabs = Colaborador.objects.filter(negocio=n, rol='ASESOR').values_list('user__first_name', 'user__last_name')
        asesores_data.append({
            'nombre': n.nombre,
            'total_asesores': len(colabs),
            'personas': [f"{fn} {ln}" for fn, ln in colabs][:10] # Solo top 10 para no saturar
        })
        
    context = {
        'admin_data_json': json.dumps(admin_data),
        'asesores_data_json': json.dumps(asesores_data)
    }
    
    return render(request, 'intranet/rrhh/organigrama.html', context)


@login_required(login_url='login')
def sincronizar_taxonomia(request):
    if not request.user.is_superuser:
        messages.error(request, "Acceso denegado.")
        return redirect('inicio')
        
    import os
    import sys
    from intranet.models.rrhh_core import Negocio, Area, Cargo
    
    # EJECUTAR EL CÃ“DIGO DE POBLACIÃ“N DIRECTAMENTE AQUÃ
    negocios_data = [
        "BAN BIF - Vigente y Dracma", "BAN BIF - Preventiva", "BAN BIF - Temprana", "BAN BIF - Castigo",
        "BBVA TARDÃAS - ExtraJudicial", "BBVA TARDÃAS - Judicial", "BBVA TARDÃAS - Castigo",
        "BBVA CONTINENTAL - CHALLENGER", "BBVA CONTINENTAL - Consumer",
        "BBVA TEMPRANAS - Particulares Vencida", "BBVA TEMPRANAS - Convenios", "BBVA TEMPRANAS - Castigo",
        "CAMPO - Campo", "CAMPO BBVA - Campo",
        "COMPARTAMOS BANCO - Vigente", "COMPARTAMOS BANCO - Grupal Liga A", "COMPARTAMOS BANCO - Castigo (Individual)",
        "FINANCIERA EFECTIVA - Xperto 1 y 2", "FINANCIERA EFECTIVA - Castigo / CASTIGO",
        "CAJA HUANCAYO", "IBK - BPE", "VOLVO", "RJ COMPARTAMOS", "RJ ADMINISTRATIVO", "Cartera Judicial"
    ]
    
    negocios_ids = []
    for neg_nombre in negocios_data:
        n, _ = Negocio.objects.get_or_create(nombre=neg_nombre)
        negocios_ids.append(n.id)
        
    areas_cargos_data = {
        "BANCOS - Sin Ã¡rea especÃ­fica": ["SUPERVISOR ESAN", "Supervisor(a) de GestiÃ³n"],
        "BANCOS - BAN BIF": ["Coordinador(a) Junior"],
        "BANCOS - BBVA CASTIGO / BVVA CASTIGO": ["Coordinador(a) Junior", "Coordinador(a) de GestiÃ³n", "Supervisora de Gestion (BBVA Castigo)"],
        "BANCOS - BBVA Convenios/Delfos": ["Supervisor(a) de GestiÃ³n"],
        "BANCOS - BBVA ExtraJudicial": ["Supervisor de Gestion"],
        "BANCOS - BBVA JUDICIAL": ["Coordinador(a) Junior"],
        "BANCOS - BBVA Particulares Vencida": ["Coordinador(a) de GestiÃ³n", "Supervisor(a) de GestiÃ³n", "Supervisora (BBVA Prev. - Particulares . Vcda. Tard)"],
        "BANCOS - COMPARTAMOS BANCO / RJ .COMPARTAMOS": ["Coordinador(a) de GestiÃ³n"],
        "BANCOS - Efectiva - IBK BPE": ["Supervisora de Gestion"],
        "BANCOS - FINANCIERA EFECTIVA": ["Supervisora (EFECTIVA - XPERTO - CASTIGO)"],
        "BANCOS - Pymes Vencida": ["JEFE DE GESTION (MAF - Vda Pymes)"],
        "BANCOS - Tempranas": ["SUBGERENCIA COMERCIAL"],
        "CAMPO - Campo": ["Coordinador(a) de Campo", "Supervisor de Gestion de Campo - SENIOR"],
        "RJ ADMINISTRATIVO - AdministraciÃ³n General": ["Jefe de AdministraciÃ³n", "Administrativo", "Asistente Administrativo", "Gerente General"],
        "RJ ADMINISTRATIVO - Ãrea Back Office": ["Back Office", "Supervisor(a) BACK OFFICE", "Digitadora"],
        "RJ ADMINISTRATIVO - Ãrea de Calidad y FormaciÃ³n": ["Jefe de Calidad y RR. HH.", "Coordinador(a) de Calidad y Formacion", "Coordinador(a) de Calidad", "Supervisor(a) de Calidad", "Asistente de Calidad y Formacion", "Asistente de Calidad"],
        "RJ ADMINISTRATIVO - Ãrea de Contabilidad y Finanzas": ["Contador", "Coordinador (a) de Contabilidad", "Asistente de Contabilidad", "Asistente Contable"],
        "RJ ADMINISTRATIVO - Ãrea Legal": ["Asesor Legal"],
        "RJ ADMINISTRATIVO - Ãrea de Mantenimiento y LogÃ­stica": ["Asistente de LogÃ­stica / Asistente de Logistica"],
        "RJ ADMINISTRATIVO - Ãrea de Recursos Humanos": ["Supervisora de RRHH", "Coordinador(a) de RRHH", "Asistente de RRHH", "MÃ©dico Ocupacional", "Coordinador de CapacitaciÃ³n"],
        "RJ ADMINISTRATIVO - Ãrea de Seguridad": ["Jefe de Seguridad", "Coordinador de Seguridad"],
        "RJ ADMINISTRATIVO - Ãrea de TecnologÃ­as de la InformaciÃ³n (Sistemas)": ["Supervisor InformÃ¡tico", "Supervisor TI", "Coordinador(a) de Sistemas", "Asistente de Sistemas", "Asistente (Monitoreo Of. San Isidro - LIMA)", "Asistente Soporte TÃ©cnico"],
        "RJ ADMINISTRATIVO - Central TelefÃ³nica": ["Central TelefÃ³nica"],
        "RJ ADMINISTRATIVO - Marketing": ["Asistente de Marketing"],
        "RJ ADMINISTRATIVO - Mandos Medios / Operativos": ["Jefe de Productividad", "Coordinador Zonal", "Supervisor(a) Zonal", "Supervisor Junior de Gestion", "Supervisor(a) de GestiÃ³n", "Asistente de GestiÃ³n"]
    }
    
    areas_activas_ids = []
    cargos_activos_ids = []
    
    for area_name, cargos in areas_cargos_data.items():
        area, _ = Area.objects.get_or_create(nombre=area_name)
        area.activa = True
        area.save()
        areas_activas_ids.append(area.id)
        
        for cargo_name in cargos:
            cargo, _ = Cargo.objects.get_or_create(area=area, nombre=cargo_name)
            cargo.activa = True
            cargo.save()
            cargos_activos_ids.append(cargo.id)
            
    # Ocultar Ã¡reas no listadas
    Area.objects.exclude(id__in=areas_activas_ids).update(activa=False)
    Cargo.objects.exclude(id__in=cargos_activos_ids).update(activa=False)
    
    messages.success(request, "Â¡TaxonomÃ­a sincronizada correctamente! Se han creado y ordenado todas las Ã¡reas, cargos y negocios (carteras).")
    return redirect('inicio')

# ==========================================
# MOTOR DE ENCUESTAS, COMUNICADOS, CALENDARIO...
# ==========================================
@login_required(login_url='login')
def encuestas_personal(request):
    perfil = getattr(request.user, 'perfil', None)

    if request.method == 'POST' and request.POST.get('enviar_encuesta') == '1' and perfil:
        encuesta = get_object_or_404(Encuesta.objects.prefetch_related('preguntas__opciones'), id=request.POST.get('encuesta_id'), activa=True)

        if RespuestaEncuesta.objects.filter(pregunta__encuesta=encuesta, colaborador=perfil).exists():
            messages.info(request, 'Ya registraste tus respuestas en este formulario.')
            return redirect('encuestas_personal')

        preguntas = list(encuesta.preguntas.all())
        errores = []
        payload = []
        respuestas_previas = {}

        for pregunta in preguntas:
            campo = f'pregunta_{pregunta.id}'
            valor = (request.POST.get(campo) or '').strip()

            if pregunta.depende_de_id:
                valor_padre = respuestas_previas.get(pregunta.depende_de_id)
                if str(valor_padre).strip().upper() != (pregunta.valor_disparador or '').strip().upper():
                    continue

            if pregunta.tipo == 'ABIERTA':
                if pregunta.obligatoria and not valor:
                    errores.append(f'La pregunta "{pregunta.texto}" es obligatoria.')
                payload.append({'pregunta': pregunta, 'valor_texto': valor or None})
                respuestas_previas[pregunta.id] = valor

            elif pregunta.tipo == 'CERRADA':
                if pregunta.obligatoria and valor not in ['SI', 'NO']:
                    errores.append(f'La pregunta "{pregunta.texto}" requiere respuesta SÃ­/No.')
                payload.append({'pregunta': pregunta, 'valor_si_no': True if valor == 'SI' else False if valor == 'NO' else None})
                respuestas_previas[pregunta.id] = valor

            elif pregunta.tipo == 'OPCION_UNICA':
                opcion_obj = pregunta.opciones.filter(id=valor).first() if valor else None
                if pregunta.obligatoria and not opcion_obj:
                    errores.append(f'La pregunta "{pregunta.texto}" requiere seleccionar una opciÃ³n.')
                payload.append({'pregunta': pregunta, 'valor_opcion': opcion_obj})
                respuestas_previas[pregunta.id] = opcion_obj.texto if opcion_obj else ''

            elif pregunta.tipo == 'ESCALA_1_5':
                numero = int(valor) if valor.isdigit() else None
                if numero is not None and (numero < 1 or numero > 5):
                    numero = None
                if pregunta.obligatoria and numero is None:
                    errores.append(f'La pregunta "{pregunta.texto}" requiere una escala del 1 al 5.')
                payload.append({'pregunta': pregunta, 'valor_numero': numero})
                respuestas_previas[pregunta.id] = str(numero) if numero is not None else ''

            elif pregunta.tipo == 'FECHA':
                fecha_valor = None
                if valor:
                    try:
                        fecha_valor = datetime.strptime(valor, '%Y-%m-%d').date()
                    except ValueError:
                        fecha_valor = None
                if pregunta.obligatoria and not fecha_valor:
                    errores.append(f'La pregunta "{pregunta.texto}" requiere una fecha vÃ¡lida.')
                payload.append({'pregunta': pregunta, 'valor_fecha': fecha_valor})
                respuestas_previas[pregunta.id] = fecha_valor.isoformat() if fecha_valor else ''

        if errores:
            for error in errores[:3]:
                messages.error(request, error)
            return redirect('encuestas_personal')

        with transaction.atomic():
            for item in payload:
                RespuestaEncuesta.objects.create(
                    pregunta=item['pregunta'],
                    colaborador=perfil,
                    sesion_id=request.session.session_key or None,
                    valor_texto=item.get('valor_texto'),
                    valor_si_no=item.get('valor_si_no'),
                    valor_opcion=item.get('valor_opcion'),
                    valor_numero=item.get('valor_numero'),
                    valor_fecha=item.get('valor_fecha'),
                )

        _notificar_respuesta_encuesta(encuesta, perfil)

        messages.success(request, 'Formulario enviado correctamente. Tus datos se asociaron automÃ¡ticamente a tu usuario.')
        return redirect('encuestas_personal')

    encuestas = Encuesta.objects.filter(activa=True).select_related('area_permitida', 'cargo_permitido', 'cartera_vinculada').prefetch_related('preguntas').order_by('-fecha_creacion')
    visibles = []
    for encuesta in encuestas:
        if perfil_coincide_segmentacion(
            perfil,
            rol=encuesta.rol_permitido,
            area=encuesta.area_permitida,
            cargo=encuesta.cargo_permitido,
            cartera=encuesta.cartera_vinculada,
            subcartera=encuesta.subcartera_vinculada,
            publico_general=encuesta.publico_general,
        ):
            visibles.append(encuesta)

    ids_visibles = [encuesta.id for encuesta in visibles]
    respuestas_personal = {}
    encuestas_respondidas = []
    encuestas_pendientes = []

    if perfil and ids_visibles:
        respuestas_personal = {
            item['pregunta__encuesta_id']: item['ultima_respuesta']
            for item in RespuestaEncuesta.objects.filter(
                colaborador=perfil,
                pregunta__encuesta_id__in=ids_visibles,
            ).values('pregunta__encuesta_id').annotate(ultima_respuesta=Max('fecha_respuesta'))
        }

    for encuesta in visibles:
        ultima_respuesta = respuestas_personal.get(encuesta.id)
        if ultima_respuesta:
            encuestas_respondidas.append((encuesta, ultima_respuesta))
        else:
            encuestas_pendientes.append(encuesta)

    encuestas_respondidas.sort(key=lambda item: item[1], reverse=True)
    total_respondidas = len(encuestas_respondidas)

    return render(request, 'intranet/comunicacion/encuestas_personal.html', {
        'encuestas': visibles,
        'encuestas_pendientes': encuestas_pendientes,
        'encuestas_respondidas': encuestas_respondidas,
        'total_visibles': len(visibles),
        'total_pendientes': len(encuestas_pendientes),
        'total_respondidas': total_respondidas,
        'perfil': perfil,
    })
@login_required(login_url='login')
@solo_directivos
def encuestas_admin(request):
    if request.method == 'POST':
        if request.POST.get('crear_encuesta') == '1':
            encuesta = Encuesta.objects.create(
                titulo=request.POST.get('titulo', '').strip(),
                descripcion=request.POST.get('descripcion', '').strip(),
                es_anonima=request.POST.get('es_anonima') == '1',
                con_puntaje=request.POST.get('con_puntaje') == '1',
                publico_general=request.POST.get('publico_general') == 'on',
                rol_permitido=request.POST.get('rol_permitido') or None,
                area_permitida=Area.objects.filter(id=request.POST.get('area_permitida')).first() if request.POST.get('area_permitida') else None,
                cargo_permitido=Cargo.objects.select_related('area').filter(id=request.POST.get('cargo_permitido')).first() if request.POST.get('cargo_permitido') else None,
                cartera_vinculada=Negocio.objects.filter(id=request.POST.get('cartera_vinculada')).first() if request.POST.get('cartera_vinculada') else None,
                subcartera_vinculada=request.POST.get('subcartera_vinculada') or None,
            )
            if encuesta.cargo_permitido and not encuesta.area_permitida and encuesta.cargo_permitido.area:
                encuesta.area_permitida = encuesta.cargo_permitido.area
                encuesta.save(update_fields=['area_permitida'])
            messages.success(request, 'Encuesta creada correctamente.')
            return redirect('encuestas_admin')

        if request.POST.get('crear_pregunta') == '1':
            encuesta = get_object_or_404(Encuesta, id=request.POST.get('encuesta_id'))
            orden = encuesta.preguntas.count() + 1
            pregunta = Pregunta.objects.create(
                encuesta=encuesta,
                texto=request.POST.get('texto', '').strip(),
                descripcion_ayuda=request.POST.get('descripcion_ayuda', '').strip(),
                tipo=request.POST.get('tipo') or 'ABIERTA',
                puntos_si=request.POST.get('puntos_si') or 0,
                obligatoria=request.POST.get('obligatoria') == 'on',
                orden=orden,
                depende_de=encuesta.preguntas.filter(id=request.POST.get('depende_de')).first() if request.POST.get('depende_de') else None,
                valor_disparador=(request.POST.get('valor_disparador') or '').strip(),
            )

            opciones_raw = (request.POST.get('opciones_texto') or '').strip()
            if pregunta.tipo == 'OPCION_UNICA' and opciones_raw:
                for idx, texto_opcion in enumerate([line.strip() for line in opciones_raw.splitlines() if line.strip()], start=1):
                    OpcionPregunta.objects.create(pregunta=pregunta, texto=texto_opcion[:180], orden=idx)

            messages.success(request, 'Pregunta agregada a la encuesta.')
            return redirect('encuestas_admin')

    encuestas = Encuesta.objects.all().select_related('area_permitida', 'cargo_permitido', 'cartera_vinculada').prefetch_related('preguntas__opciones')
    return render(request, 'intranet/admin/encuestas_admin.html', {
        'encuestas': encuestas,
        'negocios': Negocio.objects.all(),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).select_related('area').order_by('nombre'),
        'roles': Colaborador.ROLES,
        'estadisticas': {
            'total': encuestas.count(),
            'activas': encuestas.filter(activa=True).count(),
            'segmentadas': encuestas.exclude(
                publico_general=True,
                rol_permitido__isnull=True,
                area_permitida__isnull=True,
                cargo_permitido__isnull=True,
                cartera_vinculada__isnull=True,
                subcartera_vinculada__isnull=True,
            ).count(),
            'preguntas': Pregunta.objects.filter(encuesta__in=encuestas).count(),
            'respuestas': RespuestaEncuesta.objects.filter(pregunta__encuesta__in=encuestas).count(),
        },
        'encuestas_recientes': encuestas.order_by('-fecha_creacion')[:4],
    })

@login_required(login_url='login')
@solo_directivos
def crear_encuesta_view(request):
    import json
    if request.method == 'POST':
        # Procesar Survey Builder de una sola vez
        titulo = request.POST.get('titulo', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        publico_general = request.POST.get('publico_general') == 'on'
        rol_permitido = request.POST.get('rol_permitido') or None
        area_permitida_id = request.POST.get('area_permitida')
        cargo_permitido_id = request.POST.get('cargo_permitido')
        cartera_id = request.POST.get('cartera_vinculada')

        encuesta = Encuesta.objects.create(
            titulo=titulo,
            descripcion=descripcion,
            es_anonima=request.POST.get('es_anonima') == 'on',
            con_puntaje=request.POST.get('con_puntaje') == 'on',
            publico_general=publico_general,
            rol_permitido=rol_permitido,
            area_permitida=Area.objects.filter(id=area_permitida_id).first() if area_permitida_id else None,
            cargo_permitido=Cargo.objects.filter(id=cargo_permitido_id).first() if cargo_permitido_id else None,
            cartera_vinculada=Negocio.objects.filter(id=cartera_id).first() if cartera_id else None,
        )

        # Procesar JSON de preguntas
        preguntas_data = request.POST.get('preguntas_data')
        if preguntas_data:
            try:
                preguntas = json.loads(preguntas_data)
                for i, p_data in enumerate(preguntas, 1):
                    pregunta = Pregunta.objects.create(
                        encuesta=encuesta,
                        texto=p_data.get('texto', ''),
                        descripcion_ayuda=p_data.get('ayuda', ''),
                        tipo=p_data.get('tipo', 'ABIERTA'),
                        obligatoria=p_data.get('obligatoria', True),
                        orden=i
                    )
                    # Opciones si es cerrada u opcion unica
                    if pregunta.tipo in ['OPCION_UNICA', 'CERRADA'] and 'opciones' in p_data:
                        for j, op_text in enumerate(p_data['opciones'], 1):
                            if op_text.strip():
                                OpcionPregunta.objects.create(pregunta=pregunta, texto=op_text.strip(), orden=j)
            except Exception as e:
                pass # Manejo simple de error

        messages.success(request, f"Â¡Encuesta '{titulo}' creada exitosamente!")
        return redirect('encuestas_admin')

    return render(request, 'intranet/admin/encuestas_crear.html', {
        'negocios': Negocio.objects.all(),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).order_by('nombre'),
        'roles': Colaborador.ROLES,
    })
@login_required(login_url='login')
@solo_directivos
def resultados_encuesta(request, pk):
    encuesta = get_object_or_404(Encuesta.objects.prefetch_related('preguntas__opciones'), id=pk)
    preguntas = list(encuesta.preguntas.all())
    respuestas_qs = RespuestaEncuesta.objects.filter(pregunta__encuesta=encuesta).select_related('colaborador__user', 'pregunta', 'valor_opcion')

    colaboradores_respondieron = Colaborador.objects.filter(
        respuestaencuesta__pregunta__encuesta=encuesta
    ).select_related('user').distinct().order_by('user__last_name', 'user__first_name')
    total_respuestas = respuestas_qs.count()
    total_respondientes = colaboradores_respondieron.count()

    resumen = []
    for pregunta in preguntas:
        data = {
            'pregunta': pregunta,
            'total': respuestas_qs.filter(pregunta=pregunta).count(),
            'si': respuestas_qs.filter(pregunta=pregunta, valor_si_no=True).count(),
            'no': respuestas_qs.filter(pregunta=pregunta, valor_si_no=False).count(),
            'promedio': None,
            'opciones': [],
        }
        if pregunta.tipo == 'ESCALA_1_5':
            valores = [r.valor_numero for r in respuestas_qs.filter(pregunta=pregunta) if r.valor_numero is not None]
            if valores:
                data['promedio'] = round(sum(valores) / len(valores), 2)
        if pregunta.tipo == 'OPCION_UNICA':
            conteos = respuestas_qs.filter(pregunta=pregunta, valor_opcion__isnull=False).values('valor_opcion__texto').annotate(total=Count('id')).order_by('-total')
            data['opciones'] = list(conteos)
        resumen.append(data)

    return render(request, 'intranet/comunicacion/resultados_encuesta.html', {
        'encuesta': encuesta,
        'resumen': resumen,
        'colaboradores_respondieron': [] if encuesta.es_anonima else colaboradores_respondieron,
        'respondientes_totales': total_respondientes,
        'total_respuestas': total_respuestas,
    })
@login_required(login_url='login')
@solo_directivos
def exportar_encuesta(request, pk):
    encuesta = get_object_or_404(Encuesta.objects.prefetch_related('preguntas__opciones'), id=pk)
    respuestas = RespuestaEncuesta.objects.filter(pregunta__encuesta=encuesta).select_related('pregunta', 'colaborador__user', 'valor_opcion').order_by('pregunta__orden', 'fecha_respuesta')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="encuesta_{encuesta.id}_{encuesta.titulo[:40].replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Encuesta', 'Pregunta', 'Tipo', 'Respondiente', 'Usuario', 'Respuesta', 'Fecha'])

    for respuesta in respuestas:
        if encuesta.es_anonima:
            respondiente = 'AnÃ³nimo'
            usuario = ''
        else:
            respondiente = respuesta.colaborador.user.get_full_name() if respuesta.colaborador and respuesta.colaborador.user else ''
            usuario = respuesta.colaborador.user.username if respuesta.colaborador and respuesta.colaborador.user else ''

        valor = ''
        if respuesta.valor_texto not in (None, ''):
            valor = respuesta.valor_texto
        elif respuesta.valor_si_no is not None:
            valor = 'SÃ­' if respuesta.valor_si_no else 'No'
        elif respuesta.valor_opcion_id:
            valor = respuesta.valor_opcion.texto
        elif respuesta.valor_numero is not None:
            valor = str(respuesta.valor_numero)
        elif respuesta.valor_fecha:
            valor = respuesta.valor_fecha.isoformat()

        writer.writerow([
            encuesta.titulo,
            respuesta.pregunta.texto,
            respuesta.pregunta.get_tipo_display(),
            respondiente,
            usuario,
            valor,
            timezone.localtime(respuesta.fecha_respuesta).strftime('%Y-%m-%d %H:%M'),
        ])

    return response
@login_required(login_url='login')
def mensajeria(request):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil:
        return redirect('inicio')

    if request.method == 'POST' and request.POST.get('enviar_mensaje'):
        destinatarios_ids = request.POST.getlist('destinatarios')
        asunto = request.POST.get('asunto', '').strip()
        cuerpo = request.POST.get('cuerpo', '').strip()
        adjunto = request.FILES.get('adjunto')

        if not destinatarios_ids or not asunto or not cuerpo:
            messages.error(request, 'Completa destinatarios, asunto y mensaje.')
            return redirect('mensajeria')
        if not adjunto_comunicacion_valido(adjunto):
            messages.error(request, 'El adjunto no es vÃ¡lido. Usa archivos permitidos de hasta 15 MB.')
            return redirect('mensajeria')

        destinatarios = Colaborador.objects.filter(id__in=destinatarios_ids).exclude(id=perfil.id).select_related('user')
        if not destinatarios.exists():
            messages.error(request, 'Selecciona al menos un destinatario vÃ¡lido.')
            return redirect('mensajeria')

        adjunto_bytes = adjunto.read() if adjunto else None
        for destinatario in destinatarios:
            mensaje = MensajeInterno.objects.create(
                remitente=perfil,
                destinatario=destinatario,
                asunto=asunto[:200],
                cuerpo=cuerpo,
            )
            if adjunto_bytes is not None:
                mensaje.adjunto.save(adjunto.name, ContentFile(adjunto_bytes), save=True)
        messages.success(request, 'Mensaje enviado correctamente.')
        return redirect('mensajeria')

    query = request.GET.get('q', '').strip()
    recibidos = MensajeInterno.objects.filter(destinatario=perfil).select_related('remitente__user', 'destinatario__user').order_by('-fecha_envio')
    enviados = MensajeInterno.objects.filter(remitente=perfil).select_related('remitente__user', 'destinatario__user').order_by('-fecha_envio')

    if query:
        recibidos = recibidos.filter(Q(remitente__user__first_name__icontains=query) | Q(remitente__user__last_name__icontains=query) | Q(asunto__icontains=query))
        enviados = enviados.filter(Q(destinatario__user__first_name__icontains=query) | Q(destinatario__user__last_name__icontains=query) | Q(asunto__icontains=query))

    compañeros = Colaborador.objects.exclude(id=perfil.id).select_related('user').order_by('user__last_name', 'user__first_name')
    return render(request, 'intranet/comunicacion/mensajeria.html', {
        'recibidos': recibidos,
        'enviados': enviados,
        'compañeros': compañeros,
        'query': query,
    })
@login_required(login_url='login')
def leer_mensaje(request, pk):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil:
        return redirect('inicio')

    mensaje = get_object_or_404(MensajeInterno.objects.select_related('remitente__user', 'destinatario__user'), id=pk)
    if not es_participante_mensaje(mensaje, perfil):
        raise Http404('Mensaje no disponible')

    es_receptor = mensaje.destinatario_id == perfil.id
    if es_receptor and not mensaje.leido:
        mensaje.leido = True
        mensaje.save(update_fields=['leido'])

    return render(request, 'intranet/comunicacion/leer_mensaje.html', {'mensaje': mensaje, 'es_receptor': es_receptor})
@login_required(login_url='login')
def calendario(request):
    eventos = []
    for evento in EventoCalendario.objects.all().order_by('fecha_inicio'):
        eventos.append({
            'id': evento.id,
            'title': evento.titulo,
            'start': evento.fecha_inicio.isoformat() if evento.fecha_inicio else None,
            'end': evento.fecha_fin.isoformat() if evento.fecha_fin else None,
            'description': evento.descripcion or '',
        })
    return render(
        request,
        'intranet/dashboard/calendario.html',
        {'eventos_json': json.dumps(eventos), 'es_admin': getattr(request.user, 'is_staff', False) or request.user.is_superuser},
    )
@login_required(login_url='login')
def comunicados(request):
    if request.method == 'POST':
        if not usuario_es_directivo(request.user):
            raise Http404('No autorizado')
        crear_comunicado_desde_request(request)
        return redirect('comunicados')

    return render(request, 'intranet/comunicacion/comunicados.html', {
        'comunicados': Comunicado.objects.filter(activo=True).order_by('-fecha_publicacion'),
        'es_admin': usuario_es_directivo(request.user),
    })
@login_required(login_url='login')
@solo_directivos
def gestor_comunicados(request):
    if request.method == 'POST':
        crear_comunicado_desde_request(request)
        return redirect('gestor_comunicados')
    return render(request, 'intranet/comunicacion/gestor_comunicados.html', {
        'comunicados': Comunicado.objects.filter(activo=True).order_by('-fecha_publicacion')
    })
@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def eliminar_comunicado(request, pk):
    comunicado = get_object_or_404(Comunicado, id=pk)
    comunicado.delete()
    messages.success(request, 'Comunicado eliminado correctamente.')
    return redirect('gestor_comunicados')
@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def eliminar_evento(request, pk): return redirect('calendario')
@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def eliminar_candidato(request, pk): return redirect('dashboard')
@login_required(login_url='login')
@solo_directivos
def activos(request): return render(request, 'intranet/dashboard/activos.html')
@login_required(login_url='login')
def beneficios(request): return render(request, 'intranet/dashboard/beneficios.html')


# ==========================================
# ACADEMIA LMS: GESTOR Y EXÃMENES
# ==========================================
@login_required(login_url='login')
@solo_directivos
def gestor_lms(request):
    from intranet.models.rrhh_core import Negocio, Colaborador
    from intranet.models.lms import CursoInduccion, EvaluacionCurso, LeccionCurso

    if request.method == 'POST':
        if 'crear_categoria_lms' in request.POST:
            nombre_categoria = (request.POST.get('nombre_categoria') or '').strip()
            descripcion_categoria = (request.POST.get('descripcion_categoria') or '').strip()
            icono_categoria = (request.POST.get('icono_categoria') or 'bi-grid-1x2-fill').strip()
            color_categoria = (request.POST.get('color_categoria') or '#183D74').strip()

            if not nombre_categoria:
                messages.error(request, 'Debes indicar un nombre para la categoria.')
                return redirect('gestor_lms')

            categoria, created = CategoriaModuloLMS.objects.get_or_create(
                nombre=nombre_categoria,
                defaults={
                    'descripcion': descripcion_categoria,
                    'icono': icono_categoria,
                    'color': color_categoria,
                    'activa': True,
                },
            )
            if created:
                messages.success(request, f"Categoria '{nombre_categoria}' creada correctamente.")
            else:
                categoria.descripcion = descripcion_categoria or categoria.descripcion
                categoria.icono = icono_categoria or categoria.icono
                categoria.color = color_categoria or categoria.color
                categoria.activa = True
                categoria.save(update_fields=['descripcion', 'icono', 'color', 'activa'])
                messages.info(request, f"La categoria '{nombre_categoria}' ya existia. Se actualizo su configuracion.")
            return redirect('gestor_lms')

        if 'crear_curso' in request.POST:
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            portada = request.FILES.get('portada')
            portada_ok, portada_error = portada_curso_valida(portada)
            if not portada_ok:
                messages.error(request, portada_error)
                return redirect('gestor_lms')
            publico_general = request.POST.get('publico_general') == 'on'
            rol_permitido = request.POST.get('rol_permitido') or None
            area_permitida_id = request.POST.get('area_permitida')
            cargo_permitido_id = request.POST.get('cargo_permitido')
            cartera_id = request.POST.get('cartera_vinculada')
            cartera_obj = Negocio.objects.filter(id=cartera_id).first() if cartera_id else None
            categoria_text = request.POST.get('categoria') or None
            categoria_id = request.POST.get('categoria_lms')
            categoria_obj = CategoriaModuloLMS.objects.filter(id=categoria_id, activa=True).first() if categoria_id else None
            prerequisito_id = request.POST.get('prerequisito_curso')
            modulo_padre_id = request.POST.get('modulo_padre')
            nivel = request.POST.get('nivel') or 'BASICO'
            duracion_estimada_horas = request.POST.get('duracion_estimada_horas') or 1
            orden_sugerido = request.POST.get('orden_sugerido') or 1
            obligatorio = request.POST.get('obligatorio') == 'on'
            certificado_habilitado = request.POST.get('certificado_habilitado') == 'on'
            area_obj = Area.objects.filter(id=area_permitida_id).first() if area_permitida_id else None
            cargo_obj = Cargo.objects.select_related('area').filter(id=cargo_permitido_id).first() if cargo_permitido_id else None
            if cargo_obj and not area_obj and cargo_obj.area:
                area_obj = cargo_obj.area
            prerequisito_obj = CursoInduccion.objects.filter(id=prerequisito_id, tipo='ACADEMIA').first() if prerequisito_id else None
            modulo_padre_obj = CursoInduccion.objects.filter(id=modulo_padre_id, tipo='ACADEMIA').first() if modulo_padre_id else None

            CursoInduccion.objects.create(
                titulo=titulo,
                descripcion=descripcion,
                tipo='ACADEMIA', 
                publico_general=publico_general,
                rol_permitido=rol_permitido,
                area_permitida=area_obj,
                cargo_permitido=cargo_obj,
                cartera_vinculada=cartera_obj,
                subcartera_vinculada=categoria_text,
                categoria_lms=categoria_obj,
                prerequisito_curso=prerequisito_obj,
                modulo_padre=modulo_padre_obj,
                nivel=nivel,
                duracion_estimada_horas=duracion_estimada_horas,
                orden_sugerido=orden_sugerido,
                obligatorio=obligatorio,
                certificado_habilitado=certificado_habilitado,
                portada=portada,
                version=int(request.POST.get('version') or 1),
                estado_publicacion=request.POST.get('estado_publicacion') or 'PUBLICADO',
            )
            messages.success(request, f"Â¡Curso '{titulo}' creado exitosamente en la Academia LMS!")
            return redirect('gestor_lms')

        elif 'crear_leccion' in request.POST:
            curso_id = request.POST.get('curso_id')
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            url_video = request.POST.get('url_video')
            orden = request.POST.get('orden', 1)
            archivo_pdf = request.FILES.get('archivo_pdf') # Capacidad de recibir PDFs

            curso = get_object_or_404(CursoInduccion, id=curso_id)
            LeccionCurso.objects.create(
                curso=curso,
                titulo=titulo,
                descripcion=descripcion,
                url_video=url_video,
                url_simulador=request.POST.get('url_simulador') or None,
                paquete_scorm_url=request.POST.get('paquete_scorm_url') or None,
                archivo_pdf=archivo_pdf,
                orden=orden
            )
            messages.success(request, f"Â¡Clase '{titulo}' agregada correctamente al curso!")
            return redirect('gestor_lms')

        elif 'crear_evaluacion' in request.POST:
            curso_id = request.POST.get('curso_id')
            titulo = request.POST.get('titulo')
            instrucciones = request.POST.get('instrucciones', '')
            p_maximo = request.POST.get('puntaje_maximo', 20.00)
            p_aprobatorio = request.POST.get('puntaje_aprobatorio', 14.00)
            p_mostrar = request.POST.get('preguntas_a_mostrar', 10)
            aleatorio = request.POST.get('orden_aleatorio') == 'on'
            
            # Nuevos campos de gamificaciÃ³n
            tiempo_limite = request.POST.get('tiempo_limite_minutos', 0)
            puntos_premio = request.POST.get('puntos_premio', 50)
            intentos_maximos = request.POST.get('intentos_maximos') or 0
            mostrar_resultado_inmediato = request.POST.get('mostrar_resultado_inmediato') == 'on'
            permitir_revision_respuestas = request.POST.get('permitir_revision_respuestas') == 'on'
            retroalimentacion_final = request.POST.get('retroalimentacion_final', '').strip()

            curso = get_object_or_404(CursoInduccion, id=curso_id)

            if hasattr(curso, 'evaluacion'):
                messages.error(request, f"El curso '{curso.titulo}' ya tiene una evaluaciÃ³n configurada.")
            else:
                EvaluacionCurso.objects.create(
                    curso=curso, titulo=titulo, instrucciones=instrucciones,
                    puntaje_maximo=p_maximo, puntaje_aprobatorio=p_aprobatorio,
                    preguntas_a_mostrar=p_mostrar, orden_aleatorio=aleatorio,
                    tiempo_limite_minutos=tiempo_limite, puntos_premio=puntos_premio,
                    intentos_maximos=intentos_maximos,
                    mostrar_resultado_inmediato=mostrar_resultado_inmediato,
                    permitir_revision_respuestas=permitir_revision_respuestas,
                    retroalimentacion_final=retroalimentacion_final,
                )
                messages.success(request, "Â¡Examen creado! Ahora puedes subir el balotario de preguntas.")
            return redirect('gestor_lms')

    # Traemos las categorÃ­as y cursos segmentados
    categoria_filtro = request.GET.get('categoria', '')
    categorias = list(CategoriaModuloLMS.objects.filter(activa=True).order_by('nombre'))
    categorias_base = [c[0] for c in CursoInduccion.CATEGORIAS_LMS]
    cursos_sin_categoria = CursoInduccion.objects.filter(
        tipo='ACADEMIA'
    ).filter(
        Q(subcartera_vinculada__isnull=True) | Q(subcartera_vinculada='')
    ).prefetch_related('lecciones', 'evaluacion')

    cursos_qs = CursoInduccion.objects.filter(activo=True, tipo='ACADEMIA')
    if categoria_filtro:
        cursos_qs = cursos_qs.filter(
            Q(subcartera_vinculada=categoria_filtro) |
            Q(categoria_lms__nombre=categoria_filtro)
        )

    cursos_disponibles = cursos_qs.select_related('evaluacion', 'categoria_lms').prefetch_related('lecciones')

    return render(request, 'intranet/lms/gestor_lms.html', {
        'categorias': categorias,
        'categorias_base': categorias_base,
        'categoria_filtro': categoria_filtro,
        'cursos_sin_categoria': cursos_sin_categoria,
        'cursos': cursos_disponibles,
        'cursos_referencia': CursoInduccion.objects.filter(tipo='ACADEMIA').order_by('titulo'),
        'negocios': Negocio.objects.all(),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).select_related('area').order_by('nombre'),
        'roles': Colaborador.ROLES,
        'rutas': RutaInduccion.objects.all()
    })

@login_required(login_url='login')
@solo_directivos
def editar_curso_lms(request, curso_id):
    curso = get_object_or_404(CursoInduccion, id=curso_id)
    
    if request.method == 'POST':
        curso.titulo = request.POST.get('titulo')
        curso.descripcion = request.POST.get('descripcion')
        curso.publico_general = request.POST.get('publico_general') == 'on'
        curso.obligatorio = request.POST.get('obligatorio') == 'on'
        curso.certificado_habilitado = request.POST.get('certificado_habilitado') == 'on'
        curso.rol_permitido = request.POST.get('rol_permitido') or None
        curso.nivel = request.POST.get('nivel') or curso.nivel
        curso.duracion_estimada_horas = request.POST.get('duracion_estimada_horas') or curso.duracion_estimada_horas
        curso.orden_sugerido = request.POST.get('orden_sugerido') or curso.orden_sugerido
        curso.version = int(request.POST.get('version') or curso.version)
        curso.estado_publicacion = request.POST.get('estado_publicacion') or curso.estado_publicacion
        area_id = request.POST.get('area_permitida')
        cargo_id = request.POST.get('cargo_permitido')
        curso.area_permitida = Area.objects.filter(id=area_id).first() if area_id else None
        curso.cargo_permitido = Cargo.objects.select_related('area').filter(id=cargo_id).first() if cargo_id else None
        if curso.cargo_permitido and not curso.area_permitida and curso.cargo_permitido.area:
            curso.area_permitida = curso.cargo_permitido.area
        
        cartera_id = request.POST.get('cartera_vinculada')
        curso.cartera_vinculada_id = cartera_id if cartera_id else None
        categoria_lms_id = request.POST.get('categoria_lms')
        curso.categoria_lms = CategoriaModuloLMS.objects.filter(id=categoria_lms_id, activa=True).first() if categoria_lms_id else None
        curso.subcartera_vinculada = request.POST.get('categoria') or curso.subcartera_vinculada
        prerequisito_id = request.POST.get('prerequisito_curso')
        curso.prerequisito_curso = CursoInduccion.objects.filter(id=prerequisito_id, tipo='ACADEMIA').exclude(id=curso.id).first() if prerequisito_id else None
        modulo_padre_id = request.POST.get('modulo_padre')
        curso.modulo_padre = CursoInduccion.objects.filter(id=modulo_padre_id, tipo='ACADEMIA').exclude(id=curso.id).first() if modulo_padre_id else None
        
        curso.save()
        messages.success(request, f"Â¡Curso '{curso.titulo}' actualizado correctamente!")

    if request.GET.get('next') == 'curriculum':
        return redirect('curso_curriculum', pk=curso.id)
    return redirect('gestor_lms')


@login_required(login_url='login')
@solo_directivos
def duplicar_version_curso(request, curso_id):
    curso = get_object_or_404(CursoInduccion, id=curso_id)
    nueva_version = CursoInduccion.objects.create(
        titulo=curso.titulo,
        descripcion=curso.descripcion,
        tipo=curso.tipo,
        publico_general=curso.publico_general,
        rol_permitido=curso.rol_permitido,
        area_permitida=curso.area_permitida,
        cargo_permitido=curso.cargo_permitido,
        cartera_vinculada=curso.cartera_vinculada,
        subcartera_vinculada=curso.subcartera_vinculada,
        portada=curso.portada,
        categoria_lms=curso.categoria_lms,
        prerequisito_curso=curso.prerequisito_curso,
        modulo_padre=curso.modulo_padre,
        nivel=curso.nivel,
        duracion_estimada_horas=curso.duracion_estimada_horas,
        orden_sugerido=curso.orden_sugerido,
        obligatorio=curso.obligatorio,
        certificado_habilitado=curso.certificado_habilitado,
        version=curso.version + 1,
        curso_origen=curso.curso_origen or curso,
        estado_publicacion='BORRADOR',
        activo=True,
    )
    messages.success(request, f'Se creÃ³ la versiÃ³n {nueva_version.version} en borrador para {curso.titulo}.')
    return redirect(_destino_panel_por_tipo(curso.tipo))

@login_required(login_url='login')
@solo_directivos
def eliminar_curso_lms(request, curso_id):
    curso = get_object_or_404(CursoInduccion, id=curso_id)
    titulo = curso.titulo
    # Al eliminar el curso, Django borrarÃ¡ automÃ¡ticamente sus clases, exÃ¡menes y progreso (Cascade)
    curso.delete()
    messages.success(request, f"El curso '{titulo}' y todo su contenido ha sido eliminado.")
    return redirect('gestor_lms')

@login_required(login_url='login')
def academia(request):
    try:
        colaborador = request.user.perfil
    except:
        messages.error(request, "Tu usuario no tiene un perfil de trabajador asociado.")
        return redirect('inicio')

    cursos_disponibles = CursoInduccion.objects.filter(
        Q(publico_general=True) | 
        Q(rol_permitido=colaborador.rol) | 
        Q(area_permitida=colaborador.area) |
        Q(cargo_permitido=colaborador.cargo) |
        Q(cartera_vinculada=colaborador.negocio) |
        Q(subcartera_vinculada=colaborador.subcartera),
        activo=True,
        tipo='ACADEMIA',
        estado_publicacion='PUBLICADO',
    ).distinct().select_related('prerequisito_curso')

    cursos_filtrados = []
    for curso in cursos_disponibles.order_by('orden_sugerido', 'id'):
        if curso_prerequisito_cumplido(colaborador, curso):
            cursos_filtrados.append(curso)

    cursos_disponibles = cursos_filtrados

    mis_cursos = []
    for curso in cursos_disponibles:
        matricula, created = MatriculaCurso.objects.get_or_create(
            colaborador=colaborador, 
            curso=curso,
            defaults={'estado': 'PENDIENTE'}
        )
        mis_cursos.append(matricula)

    tarjetas_cursos = []
    for matricula in mis_cursos:
        curso = matricula.curso
        total_lecciones = curso.lecciones.count()
        lecciones_vistas = ProgresoLeccion.objects.filter(
            colaborador=colaborador,
            leccion__curso=curso,
            completada=True,
        ).count()
        progreso_modulo = int((lecciones_vistas / total_lecciones) * 100) if total_lecciones else 0
        evaluacion = getattr(curso, 'evaluacion', None)
        intentos_restantes = None
        if evaluacion:
            intentos_restantes = None if evaluacion.intentos_maximos == 0 else max(evaluacion.intentos_maximos - (matricula.intentos_realizados or 0), 0)

        tarjetas_cursos.append({
            'matricula': matricula,
            'curso': curso,
            'portada_url': curso.portada.url if curso.portada else None,
            'categoria': curso.categoria_lms.nombre if curso.categoria_lms else (curso.categoria or 'General'),
            'categoria_color': curso.categoria_lms.color if curso.categoria_lms else '#183D74',
            'categoria_icono': curso.categoria_lms.icono if curso.categoria_lms else 'bi-journal-bookmark-fill',
            'total_lecciones': total_lecciones,
            'lecciones_vistas': lecciones_vistas,
            'progreso_modulo': progreso_modulo,
            'tiene_examen': bool(evaluacion),
            'intentos_restantes': intentos_restantes,
        })

    total_cursos = len(mis_cursos)
    completados = sum(1 for m in mis_cursos if m.estado == 'COMPLETADO')
    porcentaje = int((completados / total_cursos) * 100) if total_cursos > 0 else 0

    return render(request, 'intranet/lms/academia.html', {
        'mis_cursos': mis_cursos,
        'tarjetas_cursos': tarjetas_cursos,
        'total': total_cursos,
        'completados': completados,
        'porcentaje': porcentaje,
    })

@login_required(login_url='login')
@solo_directivos
def importar_excel_balotario(request, evaluacion_id):
    evaluacion = get_object_or_404(EvaluacionCurso, id=evaluacion_id)
    destino = _destino_panel_por_tipo(evaluacion.curso.tipo)
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        excel_file = request.FILES['archivo_excel']
        try:
            nombre_tmp = default_storage.save(f'tmp/balotario_{request.user.id}.xlsx', ContentFile(excel_file.read()))
            wb = openpyxl.load_workbook(default_storage.open(nombre_tmp))
            cabeceras_excel = [str(celda.value).strip() for celda in wb.active[1] if celda.value is not None]
            request.session['ruta_excel_balotario'] = nombre_tmp
            request.session['evaluacion_id_temporal'] = evaluacion.id
            request.session['balotario_return_to'] = destino
            request.session.modified = True
            return render(request, 'intranet/lms/mapear_balotario.html', {'cabeceras': cabeceras_excel, 'evaluacion': evaluacion, 'return_url': destino})
        except Exception as e:
            messages.error(request, f"OcurriÃ³ un error leyendo el Excel: {str(e)}")
            return redirect(destino)
    return render(request, 'intranet/lms/subir_excel.html', {'evaluacion': evaluacion, 'return_url': destino})

@login_required(login_url='login')
@solo_directivos
def previsualizar_y_guardar_balotario(request):
    try:
        preguntas = request.session.get('balotario_temporal')
        eval_id = request.session.get('evaluacion_id_temporal')
        destino = request.session.get('balotario_return_to', 'gestor_lms')

        if not preguntas or not eval_id:
            messages.warning(request, "No hay ningÃºn balotario pendiente en memoria.")
            return redirect(destino)

        evaluacion = get_object_or_404(EvaluacionCurso, id=eval_id)

        if request.method == 'POST':
            with transaction.atomic():
                
                # --- LIMPIEZA ANTI-DUPLICADOS (Borrador automÃ¡tico) ---
                evaluacion.preguntas_balotario.all().delete()
                # ------------------------------------------------------

                ids_llegados = [k.split('_')[1] for k in request.POST if k.startswith('enunciado_')]
                if ids_llegados:
                    puntos_por_pregunta = round(evaluacion.puntaje_maximo / len(ids_llegados), 2)
                else:
                    puntos_por_pregunta = 0.00

                # 2. Guardamos formalmente
                for idx in ids_llegados:
                    enunciado_texto = request.POST.get(f'enunciado_{idx}')
                    
                    # Evitamos guardar preguntas que lleguen vacÃ­as
                    if not enunciado_texto or not enunciado_texto.strip():
                        continue

                    nueva_pregunta = PreguntaEvaluacion.objects.create(
                        evaluacion=evaluacion,
                        enunciado=enunciado_texto.strip(),
                        puntos=puntos_por_pregunta
                    )

                    opcion_correcta_id = request.POST.get(f'correcta_{idx}')
                    textos_agregados = set()  # <--- FILTRO MÃGICO: Evita duplicar textos en la misma pregunta
                    
                    for i in range(1, 5):
                        txt_alt = request.POST.get(f'alt{i}_{idx}')
                        
                        # Si la caja estaba vacÃ­a (ej: V o F de solo 2 opciones), la ignoramos por completo
                        if txt_alt and txt_alt.strip():
                            texto_limpio = txt_alt.strip()
                            texto_lower = texto_limpio.lower()
                            
                            # Si no hemos guardado esta opciÃ³n antes, la agregamos
                            if texto_lower not in textos_agregados:
                                es_correcta = (str(i) == str(opcion_correcta_id))
                                OpcionRespuesta.objects.create(pregunta=nueva_pregunta, texto=texto_limpio, es_correcta=es_correcta)
                                textos_agregados.add(texto_lower)

            del request.session['balotario_temporal']
            del request.session['evaluacion_id_temporal']
            request.session.pop('balotario_return_to', None)

            messages.success(request, "Â¡Balotario mapeado e inyectado con Ã©xito!")
            return redirect(destino)

        return render(request, 'intranet/lms/previsualizar_balotario.html', {'preguntas': preguntas, 'evaluacion': evaluacion, 'return_url': destino})

    except Exception as e:
        messages.error(request, "OcurriÃ³ un error procesando el balotario. Contacte a soporte.")
        return redirect(request.session.get('balotario_return_to', 'gestor_lms'))

@login_required(login_url='login')
def rendir_evaluacion(request, matricula_id):
    perfil = request.user.perfil
    matricula = get_object_or_404(MatriculaCurso, id=matricula_id, colaborador=perfil)
    
    if not hasattr(matricula.curso, 'evaluacion'):
        messages.error(request, "Este curso aÃºn no tiene un examen configurado.")
        return redirect('academia')
        
    evaluacion = matricula.curso.evaluacion

    if matricula.estado == 'COMPLETADO':
        messages.info(request, f"Ya rendiste este examen. Tu nota final fue: {matricula.nota_obtenida}")
        return redirect('academia')

    if matricula.estado == 'REPROBADO':
        if evaluacion.intentos_maximos and matricula.intentos_realizados >= evaluacion.intentos_maximos:
            messages.error(request, "Alcanzaste el limite de intentos permitidos para este examen.")
            return redirect('detalle_curso', curso_id=matricula.curso.id)
        RespuestaColaborador.objects.filter(matricula=matricula).delete()
        matricula.estado = 'PENDIENTE'
        matricula.save(update_fields=['estado'])

    if request.method == 'POST':
        nota_final = 0.00
        preguntas_ids = request.POST.getlist('preguntas_mostradas')
        preguntas_evaluadas = PreguntaEvaluacion.objects.filter(id__in=preguntas_ids).prefetch_related('alternativas')

        with transaction.atomic():
            for pregunta in preguntas_evaluadas:
                opcion_marcada_id = request.POST.get(f'pregunta_{pregunta.id}')
                puntos_ganados = 0.00
                es_correcta = False
                opcion_obj = None

                if opcion_marcada_id:
                    opcion_obj = pregunta.alternativas.filter(id=opcion_marcada_id).first()
                    if opcion_obj and opcion_obj.es_correcta:
                        puntos_ganados = float(pregunta.puntos)
                        es_correcta = True
                        nota_final += puntos_ganados

                respuesta_registro = RespuestaColaborador.objects.create(
                    matricula=matricula, pregunta=pregunta, es_correcta=es_correcta, puntos_obtenidos=puntos_ganados
                )
                if opcion_obj:
                    respuesta_registro.opciones_marcadas.add(opcion_obj)

            matricula.nota_obtenida = nota_final
            matricula.fecha_finalizacion = timezone.now()
            matricula.intentos_realizados = (matricula.intentos_realizados or 0) + 1
            
            # --- EVALUACIÃ“N DE APROBACIÃ“N Y ASIGNACIÃ“N DE PUNTOS ---
            if nota_final >= float(evaluacion.puntaje_aprobatorio):
                matricula.estado = 'COMPLETADO'
                if evaluacion.mostrar_resultado_inmediato:
                    messages.success(request, f"Â¡Felicidades! Has aprobado el examen con {nota_final} puntos.")
                else:
                    messages.success(request, "Examen enviado correctamente. Tu resultado sera revisado por el equipo de formacion.")
                
                # Â¡MAGIA DE GAMIFICACIÃ“N! Sumamos los puntos del examen al perfil del usuario
                perfil.puntos_lms = getattr(perfil, 'puntos_lms', 0) + evaluacion.puntos_premio
                perfil.save()
                matricula.certificado_codigo = generar_codigo_certificado(matricula)
                matricula.certificado_emitido_en = timezone.now()
                if not matricula.certificado_vigente_hasta:
                    matricula.certificado_vigente_hasta = date.today().replace(year=date.today().year + 1)
                Notificacion.objects.create(
                    usuario=request.user,
                    tipo='ALERTA',
                    titulo='Certificado disponible',
                    detalle=f'Ya puedes descargar tu certificado del curso {matricula.curso.titulo}.',
                    url_destino=f'/academia/certificado/{matricula.id}/',
                )
            else:
                matricula.estado = 'REPROBADO'
                intentos_restantes = 'ilimitados' if evaluacion.intentos_maximos == 0 else max(evaluacion.intentos_maximos - matricula.intentos_realizados, 0)
                messages.error(request, f"No alcanzaste la nota minima. Obtuviste {nota_final} puntos. Intentos restantes: {intentos_restantes}.")
            
            matricula.save()
            return redirect('academia')

    if matricula.estado != 'EVALUANDO':
        matricula.estado = 'EVALUANDO'
        matricula.save()

    if evaluacion.orden_aleatorio:
        total_objetivo = evaluacion.preguntas_a_mostrar
        preguntas_pool = evaluacion.preguntas_balotario.filter(activa=True)
        basicas = list(preguntas_pool.filter(dificultad='BASICO'))
        intermedias = list(preguntas_pool.filter(dificultad='INTERMEDIO'))
        avanzadas = list(preguntas_pool.filter(dificultad='AVANZADO'))

        random.shuffle(basicas)
        random.shuffle(intermedias)
        random.shuffle(avanzadas)

        cupo = max(total_objetivo // 3, 1)
        preguntas_seleccionadas = basicas[:cupo] + intermedias[:cupo] + avanzadas[:cupo]
        resto_pool = [p for p in list(preguntas_pool) if p not in preguntas_seleccionadas]
        random.shuffle(resto_pool)
        if len(preguntas_seleccionadas) < total_objetivo:
            preguntas_seleccionadas.extend(resto_pool[:total_objetivo - len(preguntas_seleccionadas)])
        random.shuffle(preguntas_seleccionadas)
        preguntas_qs = preguntas_seleccionadas[:total_objetivo]
    else:
        preguntas_qs = evaluacion.preguntas_balotario.filter(activa=True).order_by('id')[:evaluacion.preguntas_a_mostrar]

    # --- EL CHOCOLATEO DE OPCIONES ---
    preguntas = list(preguntas_qs)
    
    for p in preguntas:
        opciones = list(p.alternativas.all())
        random.shuffle(opciones)
        # Guardamos la lista barajada para usarla en el HTML
        p.opciones_mezcladas = opciones

    return render(request, 'intranet/lms/rendir_examen.html', {'matricula': matricula, 'evaluacion': evaluacion, 'preguntas': preguntas})

@login_required(login_url='login')
@solo_directivos
def resultados_evaluacion(request, evaluacion_id):
    evaluacion = get_object_or_404(EvaluacionCurso, id=evaluacion_id)
    matriculas = MatriculaCurso.objects.filter(curso=evaluacion.curso).select_related('colaborador__user')
    
    completados = matriculas.filter(estado__in=['COMPLETADO', 'REPROBADO'])
    total_completados = completados.count()
    
    promedio = 0
    if total_completados > 0:
        promedio = sum(m.nota_obtenida for m in completados) / total_completados
        
    aprobados = matriculas.filter(estado='COMPLETADO').count()
    reprobados = matriculas.filter(estado='REPROBADO').count()

   # Preparamos los datos matemÃ¡ticos para los grÃ¡ficos circulares
    graficos = []
    for pregunta in evaluacion.preguntas_balotario.filter(activa=True):
        opciones = pregunta.alternativas.all()
        labels = [op.texto for op in opciones]
        data = []
        for op in opciones:
            cantidad_votos = RespuestaColaborador.objects.filter(pregunta=pregunta, opciones_marcadas=op).count()
            data.append(cantidad_votos)
        
        graficos.append({
            'id': pregunta.id,
            'enunciado': pregunta.enunciado,
            'labels': labels, # <-- Lo dejamos como una lista normal de Python
            'data': data      # <-- Lo dejamos como una lista normal de Python
        })

    return render(request, 'intranet/lms/resultados_evaluacion.html', {
        'evaluacion': evaluacion,
        'matriculas': matriculas,
        'promedio': round(promedio, 2),
        'aprobados': aprobados,
        'reprobados': reprobados,
        'total_completados': total_completados,
        'graficos': graficos
    })

@login_required(login_url='login')
@solo_directivos
def reabrir_examen(request, matricula_id):
    matricula = get_object_or_404(MatriculaCurso, id=matricula_id)
    evaluacion_id = matricula.curso.evaluacion.id
    
    # Reseteamos los valores para que pueda darlo de nuevo
    matricula.estado = 'PENDIENTE'
    matricula.nota_obtenida = 0.00
    matricula.fecha_finalizacion = None
    matricula.save()
    
    # Destruimos sus respuestas anteriores para dejar la hoja en blanco
    RespuestaColaborador.objects.filter(matricula=matricula).delete()
    
    messages.success(request, f"Examen reabierto exitosamente para {matricula.colaborador.user.first_name}.")
    return redirect('resultados_evaluacion', evaluacion_id=evaluacion_id)

@login_required(login_url='login')
@solo_directivos
def exportar_resultados_lms(request, evaluacion_id):
    evaluacion = get_object_or_404(EvaluacionCurso, id=evaluacion_id)
    matriculas = MatriculaCurso.objects.filter(curso=evaluacion.curso, estado__in=['COMPLETADO', 'REPROBADO'])
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados Examen"
    
    # Cabeceras Base
    headers = ['Colaborador', 'DNI', 'Nota Obtenida', 'Estado', 'Fecha Examen']
    preguntas = evaluacion.preguntas_balotario.filter(activa=True).order_by('id')
    
    # Agregamos las preguntas como cabeceras en el Excel
    for p in preguntas:
        headers.append(p.enunciado)
        
    ws.append(headers)
    
    # Llenamos la data de los colaboradores
    for m in matriculas:
        row = [
            f"{m.colaborador.user.first_name} {m.colaborador.user.last_name}",
            m.colaborador.dni,
            float(m.nota_obtenida),
            m.estado,
            m.fecha_finalizacion.strftime('%Y-%m-%d %H:%M') if m.fecha_finalizacion else ''
        ]
        
        # Buscamos quÃ© marcÃ³ exactamente en cada pregunta
        for p in preguntas:
            resp = RespuestaColaborador.objects.filter(matricula=m, pregunta=p).first()
            if resp and resp.opciones_marcadas.exists():
                marcada = resp.opciones_marcadas.first().texto
                row.append(marcada)
            else:
                row.append("-")
        ws.append(row)
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Resultados_{evaluacion.titulo}.xlsx"'
    wb.save(response)
    return response

@login_required(login_url='login')
def generar_certificado(request, matricula_id):
    perfil = request.user.perfil
    matricula = get_object_or_404(MatriculaCurso, id=matricula_id, colaborador=perfil, estado='COMPLETADO')
    if not matricula.certificado_codigo:
        matricula.certificado_codigo = generar_codigo_certificado(matricula)
        matricula.certificado_emitido_en = timezone.now()
        if not matricula.certificado_vigente_hasta:
            matricula.certificado_vigente_hasta = date.today().replace(year=date.today().year + 1)
        matricula.save(update_fields=['certificado_codigo', 'certificado_emitido_en', 'certificado_vigente_hasta'])

    verificacion_url = f"/academia/certificado/verificar/{matricula.certificado_codigo}/"
    qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=180x180&data={request.build_absolute_uri(verificacion_url)}"
    
    return render(request, 'intranet/lms/certificado.html', {
        'matricula': matricula,
        'fecha_emision': date.today(),
        'verificacion_url': verificacion_url,
        'qr_url': qr_url,
    })


def verificar_certificado(request, codigo):
    matricula = MatriculaCurso.objects.filter(certificado_codigo=codigo, estado='COMPLETADO').select_related('colaborador__user', 'curso').first()
    if not matricula:
        return render(request, 'intranet/lms/verificar_certificado.html', {'valido': False, 'codigo': codigo})

    vigente = True
    if matricula.certificado_vigente_hasta and matricula.certificado_vigente_hasta < date.today():
        vigente = False

    return render(request, 'intranet/lms/verificar_certificado.html', {
        'valido': True,
        'vigente': vigente,
        'matricula': matricula,
    })

# ==========================================
# RUTAS DE APRENDIZAJE: CLASES Y VIDEOS
# ==========================================
@login_required(login_url='login')
def detalle_curso(request, curso_id):
    perfil = request.user.perfil
    curso = get_object_or_404(CursoInduccion, id=curso_id)
    matricula = get_object_or_404(MatriculaCurso, curso=curso, colaborador=perfil)
    
    # Extraemos todas las lecciones ordenadas
    lecciones = curso.lecciones.all().order_by('orden')
    
    # Revisamos cuÃ¡les lecciones ya aprobÃ³ este asesor
    lecciones_completadas = ProgresoLeccion.objects.filter(
        colaborador=perfil, completada=True
    ).values_list('leccion_id', flat=True)
    
    # Calculamos el % de progreso interno del curso
    porcentaje_progreso = 0
    if lecciones.count() > 0:
        porcentaje_progreso = int((len(lecciones_completadas) / lecciones.count()) * 100)
        
    # EL CANDADO INFALIBLE: El examen se desbloquea SOLO si vio todas las clases
    examen_desbloqueado = (lecciones.count() == len(lecciones_completadas))
    
    # Si el curso no tiene clases (solo examen), lo dejamos libre
    if lecciones.count() == 0:
        examen_desbloqueado = True

    evaluacion = getattr(curso, 'evaluacion', None)
    intentos_maximos = evaluacion.intentos_maximos if evaluacion else 1
    intentos_realizados = matricula.intentos_realizados or 0
    intentos_restantes = None if intentos_maximos == 0 else max(intentos_maximos - intentos_realizados, 0)

    return render(request, 'intranet/lms/detalle_curso.html', {
        'curso': curso,
        'matricula': matricula,
        'lecciones': lecciones,
        'completadas': lecciones_completadas,
        'progreso_lecciones': porcentaje_progreso,
        'examen_desbloqueado': examen_desbloqueado,
        'evaluacion': evaluacion,
        'intentos_maximos': intentos_maximos,
        'intentos_realizados': intentos_realizados,
        'intentos_restantes': intentos_restantes,
    })

@login_required(login_url='login')
def ver_leccion(request, leccion_id):
    perfil = getattr(request.user, 'perfil', None)
    leccion = get_object_or_404(LeccionCurso, id=leccion_id)
    
    # VALIDACIÃ“N IDOR
    if not usuario_es_directivo(request.user):
        if not perfil or not MatriculaCurso.objects.filter(colaborador=perfil, curso=leccion.curso).exists():
            raise Http404("LecciÃ³n no disponible")
            
    # Verificamos si ya la habÃ­a marcado como completada antes
    ya_completada = ProgresoLeccion.objects.filter(colaborador=perfil, leccion=leccion, completada=True).exists()
    
    return render(request, 'intranet/lms/ver_leccion.html', {
        'leccion': leccion,
        'curso': leccion.curso,
        'ya_completada': ya_completada
    })


@login_required(login_url='login')
def ver_leccion_pdf(request, leccion_id):
    leccion = get_object_or_404(LeccionCurso, id=leccion_id)

    if not leccion.archivo_pdf:
        raise Http404("Archivo no disponible")

    if not usuario_es_directivo(request.user):
        perfil = getattr(request.user, 'perfil', None)
        if not perfil or not MatriculaCurso.objects.filter(colaborador=perfil, curso=leccion.curso).exists():
            raise Http404("Archivo no disponible")

    return build_storage_response(leccion.archivo_pdf)


@login_required(login_url='login')
def ver_adjunto_mensaje(request, pk):
    perfil = getattr(request.user, 'perfil', None)
    if not perfil:
        raise Http404('Archivo no disponible')

    mensaje = get_object_or_404(MensajeInterno, id=pk)
    if not es_participante_mensaje(mensaje, perfil):
        raise Http404('Archivo no disponible')

    return build_storage_response(mensaje.adjunto)


@login_required(login_url='login')
def ver_adjunto_comunicado(request, pk):
    comunicado = get_object_or_404(Comunicado, id=pk, activo=True)
    return build_storage_response(comunicado.adjunto)

@login_required(login_url='login')
@require_http_methods(["POST"])
def completar_leccion(request, leccion_id):
    perfil = getattr(request.user, 'perfil', None)
    leccion = get_object_or_404(LeccionCurso, id=leccion_id)
    
    # VALIDACIÃ“N IDOR
    if not perfil or not MatriculaCurso.objects.filter(colaborador=perfil, curso=leccion.curso).exists():
        raise Http404("No autorizado")
        
    # Registramos que vio el video/PDF de forma indestructible
    ProgresoLeccion.objects.get_or_create(
        colaborador=perfil, 
        leccion=leccion, 
        defaults={'completada': True}
    )
    
    messages.success(request, f"Â¡Excelente! Has completado la clase: '{leccion.titulo}'")
    return redirect('detalle_curso', curso_id=leccion.curso.id)

# ==========================================
# DASHBOARD GENERAL DE RESULTADOS LMS
# ==========================================
@login_required(login_url='login')
@solo_directivos
def dashboard_resultados(request):
    # 1. MÃ©tricas generales de la plataforma
    total_cursos = CursoInduccion.objects.count()
    total_examenes = EvaluacionCurso.objects.count()
    
    # 2. MÃ©tricas de los usuarios (MatriculaCurso guarda las notas y estados)
    evaluaciones_rendidas = MatriculaCurso.objects.filter(estado__in=['COMPLETADO', 'REPROBADO'])
    total_rendidas = evaluaciones_rendidas.count()
    
    aprobados = evaluaciones_rendidas.filter(estado='COMPLETADO').count()
    reprobados = evaluaciones_rendidas.filter(estado='REPROBADO').count()
    
    # 3. Promedio global de la empresa
    promedio_dict = evaluaciones_rendidas.aggregate(promedio=Avg('nota_obtenida'))
    nota_promedio = promedio_dict['promedio'] or 0.00

    rendimiento_area = list(
        evaluaciones_rendidas.values('colaborador__area__nombre')
        .annotate(total=Count('id'), promedio=Avg('nota_obtenida'))
        .order_by('-total')[:8]
    )
    rendimiento_cargo = list(
        evaluaciones_rendidas.values('colaborador__cargo__nombre')
        .annotate(total=Count('id'), promedio=Avg('nota_obtenida'))
        .order_by('-total')[:8]
    )
    modulos_retrasados = MatriculaCurso.objects.filter(
        estado__in=['PENDIENTE', 'EN_CURSO', 'REPROBADO'],
        fecha_limite__isnull=False,
        fecha_limite__lt=date.today(),
    ).select_related('colaborador__user', 'curso').order_by('fecha_limite')[:10]

    context = {
        'total_cursos': total_cursos,
        'total_examenes_creados': total_examenes,
        'total_rendidas': total_rendidas,
        'aprobados': aprobados,
        'reprobados': reprobados,
        'nota_promedio': round(nota_promedio, 2),
        'rendimiento_area': rendimiento_area,
        'rendimiento_cargo': rendimiento_cargo,
        'modulos_retrasados': modulos_retrasados,
    }
    
    return render(request, 'intranet/lms/dashboard_resultados.html', context)

@login_required(login_url='login')
@solo_directivos
def crear_curso_avanzado(request, curso_id=None):
    from intranet.models.rrhh_core import Negocio, Colaborador
    from intranet.models.lms import CursoInduccion, LeccionCurso, EvaluacionLMS
    
    curso = None
    if curso_id:
        curso = get_object_or_404(CursoInduccion, id=curso_id)

    if request.method == 'POST':
        paso = request.POST.get('paso')

        if paso == '1':
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            portada = request.FILES.get('portada')
            
            if portada:
                portada_ok, portada_error = portada_curso_valida(portada)
                if not portada_ok:
                    messages.error(request, portada_error)
                    return redirect('crear_curso_avanzado' if not curso else 'editar_curso_avanzado', curso_id=curso.id if curso else None)
            
            categoria_text = request.POST.get('categoria') or None
            categoria_lms_id = request.POST.get('categoria_lms')
            categoria_lms = CategoriaModuloLMS.objects.filter(id=categoria_lms_id, activa=True).first() if categoria_lms_id else None
            nivel = request.POST.get('nivel') or 'BASICO'
            duracion_estimada_horas = request.POST.get('duracion_estimada_horas') or 1
            
            publico_general = request.POST.get('publico_general') == 'on'
            rol_permitido = request.POST.get('rol_permitido') or None
            cartera_id = request.POST.get('cartera_vinculada')
            cartera_obj = Negocio.objects.filter(id=cartera_id).first() if cartera_id else None

            if curso:
                curso.titulo = titulo
                curso.descripcion = descripcion
                if portada: curso.portada = portada
                curso.subcartera_vinculada = categoria_text
                curso.categoria_lms = categoria_lms
                curso.nivel = nivel
                curso.duracion_estimada_horas = duracion_estimada_horas
                curso.publico_general = publico_general
                curso.rol_permitido = rol_permitido
                curso.cartera_vinculada = cartera_obj
                curso.save()
            else:
                curso = CursoInduccion.objects.create(
                    titulo=titulo,
                    descripcion=descripcion,
                    tipo='ACADEMIA', 
                    portada=portada,
                    subcartera_vinculada=categoria_text,
                    categoria_lms=categoria_lms,
                    nivel=nivel,
                    duracion_estimada_horas=duracion_estimada_horas,
                    publico_general=publico_general,
                    rol_permitido=rol_permitido,
                    cartera_vinculada=cartera_obj,
                    estado_publicacion='BORRADOR'
                )
            return redirect(f"/intranet/lms/editar-curso/{curso.id}/?step=2")

        elif paso == '3':
            # Guardar evaluaciÃ³n
            if curso:
                if curso.lecciones.count() == 0:
                    messages.error(request, "Debes crear al menos una clase antes de configurar la evaluaciÃ³n.")
                    return redirect(f"/intranet/lms/editar-curso/{curso.id}/?step=2")

                titulo_eval = request.POST.get('titulo')
                instrucciones = request.POST.get('instrucciones', '')
                puntaje_maximo = request.POST.get('puntaje_maximo', 20.00)
                puntaje_aprobatorio = request.POST.get('puntaje_aprobatorio', 14.00)
                tiempo_limite_minutos = request.POST.get('tiempo_limite_minutos', 0)
                puntos_premio = request.POST.get('puntos_premio', 50)
                preguntas_a_mostrar = request.POST.get('preguntas_a_mostrar', 10)
                intentos_maximos = request.POST.get('intentos_maximos', 1)
                
                if hasattr(curso, 'evaluacion'):
                    ev = curso.evaluacion
                    ev.titulo = titulo_eval
                    ev.instrucciones = instrucciones
                    ev.puntaje_maximo = puntaje_maximo
                    ev.puntaje_aprobatorio = puntaje_aprobatorio
                    ev.tiempo_limite_minutos = tiempo_limite_minutos
                    ev.puntos_premio = puntos_premio
                    ev.preguntas_a_mostrar = preguntas_a_mostrar
                    ev.intentos_maximos = intentos_maximos
                    ev.save()
                else:
                    EvaluacionCurso.objects.create(
                        curso=curso,
                        titulo=titulo_eval,
                        instrucciones=instrucciones,
                        puntaje_maximo=puntaje_maximo,
                        puntaje_aprobatorio=puntaje_aprobatorio,
                        tiempo_limite_minutos=tiempo_limite_minutos,
                        puntos_premio=puntos_premio,
                        preguntas_a_mostrar=preguntas_a_mostrar,
                        intentos_maximos=intentos_maximos
                    )
            return redirect(f"/intranet/lms/editar-curso/{curso.id}/?step=4")

        elif paso == '4':
            # Publicar Curso
            if curso:
                if curso.lecciones.count() == 0:
                    messages.error(request, "No puedes publicar sin clases. Agrega al menos una clase en el paso 2.")
                    return redirect(f"/intranet/lms/editar-curso/{curso.id}/?step=2")
                if not hasattr(curso, 'evaluacion'):
                    messages.error(request, "No puedes publicar sin evaluaciÃ³n. Configura el paso 3.")
                    return redirect(f"/intranet/lms/editar-curso/{curso.id}/?step=3")

                curso.estado_publicacion = 'PUBLICADO'
                curso.save()
                messages.success(request, f"Â¡Curso '{curso.titulo}' publicado exitosamente!")
                return redirect('dashboard_resultados')

    context = {
        'curso': curso,
        'categorias': [c[0] for c in CursoInduccion.CATEGORIAS_LMS],
        'categorias_lms': CategoriaModuloLMS.objects.filter(activa=True).order_by('nombre'),
        'negocios': Negocio.objects.all(),
        'roles': Colaborador.ROLES,
        'lecciones_count': curso.lecciones.count() if curso else 0,
        'tiene_evaluacion': bool(getattr(curso, 'evaluacion', None)) if curso else False,
    }
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        pass # PodrÃ­amos devolver JSON parciales aquÃ­ si lo deseamos

    return render(request, 'intranet/lms/crear_curso_stepper.html', context)

@login_required(login_url='login')
@solo_directivos
def api_gestionar_lecciones(request, curso_id):
    import json
    from django.http import JsonResponse
    from intranet.models.lms import CursoInduccion, LeccionCurso
    
    curso = get_object_or_404(CursoInduccion, id=curso_id)

    if request.method == 'GET':
        lecciones = curso.lecciones.all().order_by('orden')
        data = [{
            'id': l.id,
            'titulo': l.titulo,
            'descripcion': l.descripcion,
            'url_video': l.url_video,
            'url_presentacion_canva': l.url_presentacion_canva,
            'url_simulador': l.url_simulador,
            'paquete_scorm_url': l.paquete_scorm_url,
            'tiene_pdf': bool(l.archivo_pdf),
            'orden': l.orden
        } for l in lecciones]
        return JsonResponse({'lecciones': data})
        
    elif request.method == 'POST':
        # Agregar leccion
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion')
        url_video = request.POST.get('url_video')
        url_presentacion_canva = request.POST.get('url_presentacion_canva')
        url_simulador = request.POST.get('url_simulador')
        paquete_scorm_url = request.POST.get('paquete_scorm_url')
        archivo_pdf = request.FILES.get('archivo_pdf')

        if not _leccion_tiene_contenido_minimo(url_video, archivo_pdf):
            return JsonResponse({'status': 'error', 'message': 'Cada clase debe tener al menos un video o un PDF.'}, status=400)

        orden = curso.lecciones.count() + 1

        leccion = LeccionCurso.objects.create(
            curso=curso,
            titulo=titulo,
            descripcion=descripcion,
            url_video=url_video,
            url_presentacion_canva=url_presentacion_canva,
            url_simulador=url_simulador,
            paquete_scorm_url=paquete_scorm_url,
            archivo_pdf=archivo_pdf,
            orden=orden
        )
        return JsonResponse({'status': 'ok', 'id': leccion.id, 'titulo': leccion.titulo})

    elif request.method == 'PUT':
        try:
            body = json.loads(request.body)

            order_ids = body.get('order_ids')
            if isinstance(order_ids, list) and order_ids:
                with transaction.atomic():
                    for idx, lid in enumerate(order_ids, start=1):
                        curso.lecciones.filter(id=lid).update(orden=idx)
                return JsonResponse({'status': 'ok'})

            leccion_id = body.get('leccion_id')
            leccion = get_object_or_404(LeccionCurso, id=leccion_id, curso=curso)

            movimiento = body.get('move')
            if movimiento in ('up', 'down'):
                if movimiento == 'up':
                    vecina = curso.lecciones.filter(orden__lt=leccion.orden).order_by('-orden').first()
                else:
                    vecina = curso.lecciones.filter(orden__gt=leccion.orden).order_by('orden').first()

                if vecina:
                    orden_actual = leccion.orden
                    leccion.orden = vecina.orden
                    vecina.orden = orden_actual
                    leccion.save(update_fields=['orden'])
                    vecina.save(update_fields=['orden'])
                return JsonResponse({'status': 'ok'})

            leccion.titulo = body.get('titulo', leccion.titulo)
            leccion.descripcion = body.get('descripcion', leccion.descripcion)
            leccion.url_video = body.get('url_video', leccion.url_video)
            leccion.url_presentacion_canva = body.get('url_presentacion_canva', leccion.url_presentacion_canva)
            leccion.url_simulador = body.get('url_simulador') or None
            leccion.paquete_scorm_url = body.get('paquete_scorm_url') or None
            if body.get('orden') is not None:
                leccion.orden = int(body.get('orden') or leccion.orden)

            if not _leccion_tiene_contenido_minimo(leccion.url_video, leccion.archivo_pdf):
                return JsonResponse({'status': 'error', 'message': 'Cada clase debe tener al menos un video o un PDF.'}, status=400)

            leccion.save()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

    elif request.method == 'DELETE':
        try:
            body = json.loads(request.body)
            leccion_id = body.get('leccion_id')
            leccion = get_object_or_404(LeccionCurso, id=leccion_id, curso=curso)
            leccion.delete()
            return JsonResponse({'status': 'ok'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@login_required(login_url='login')
@solo_directivos
def crear_curso_induccion(request, curso_id=None):
    from intranet.models.lms import CursoInduccion, EvaluacionCurso
    
    curso = None
    if curso_id:
        curso = get_object_or_404(CursoInduccion, id=curso_id)

    if request.method == 'POST':
        paso = request.POST.get('paso')

        if paso == '1':
            titulo = request.POST.get('titulo')
            descripcion = request.POST.get('descripcion')
            portada = request.FILES.get('portada')
            
            duracion_estimada_horas = request.POST.get('duracion_estimada_horas') or 1
            
            if curso:
                curso.titulo = titulo
                curso.descripcion = descripcion
                if portada: curso.portada = portada
                curso.duracion_estimada_horas = duracion_estimada_horas
                curso.save()
            else:
                curso = CursoInduccion.objects.create(
                    titulo=titulo,
                    descripcion=descripcion,
                    tipo='INDUCCION', 
                    portada=portada,
                    duracion_estimada_horas=duracion_estimada_horas,
                    publico_general=False,
                    estado_publicacion='BORRADOR'
                )
            return redirect(f"/intranet/induccion/modulo/editar/{curso.id}/?step=2")

        elif paso == '3':
            # Guardar evaluaciÃ³n
            if curso:
                if curso.lecciones.count() == 0:
                    messages.error(request, "Debes crear al menos una clase antes de configurar la evaluaciÃ³n.")
                    return redirect(f"/intranet/induccion/modulo/editar/{curso.id}/?step=2")

                titulo_eval = request.POST.get('titulo')
                instrucciones = request.POST.get('instrucciones', '')
                puntaje_maximo = request.POST.get('puntaje_maximo', 20.00)
                puntaje_aprobatorio = request.POST.get('puntaje_aprobatorio', 14.00)
                tiempo_limite_minutos = request.POST.get('tiempo_limite_minutos', 0)
                puntos_premio = request.POST.get('puntos_premio', 50)
                preguntas_a_mostrar = request.POST.get('preguntas_a_mostrar', 10)
                intentos_maximos = request.POST.get('intentos_maximos', 1)
                
                if hasattr(curso, 'evaluacion'):
                    ev = curso.evaluacion
                    ev.titulo = titulo_eval
                    ev.instrucciones = instrucciones
                    ev.puntaje_maximo = puntaje_maximo
                    ev.puntaje_aprobatorio = puntaje_aprobatorio
                    ev.tiempo_limite_minutos = tiempo_limite_minutos
                    ev.puntos_premio = puntos_premio
                    ev.preguntas_a_mostrar = preguntas_a_mostrar
                    ev.intentos_maximos = intentos_maximos
                    ev.save()
                else:
                    EvaluacionCurso.objects.create(
                        curso=curso,
                        titulo=titulo_eval,
                        instrucciones=instrucciones,
                        puntaje_maximo=puntaje_maximo,
                        puntaje_aprobatorio=puntaje_aprobatorio,
                        tiempo_limite_minutos=tiempo_limite_minutos,
                        puntos_premio=puntos_premio,
                        preguntas_a_mostrar=preguntas_a_mostrar,
                        intentos_maximos=intentos_maximos
                    )
            return redirect(f"/intranet/induccion/modulo/editar/{curso.id}/?step=4")

        elif paso == '4':
            if curso:
                if curso.lecciones.count() == 0:
                    messages.error(request, "No puedes publicar sin clases. Agrega al menos una clase en el paso 2.")
                    return redirect(f"/intranet/induccion/modulo/editar/{curso.id}/?step=2")
                if not hasattr(curso, 'evaluacion'):
                    messages.error(request, "No puedes publicar sin evaluaciÃ³n. Configura el paso 3.")
                    return redirect(f"/intranet/induccion/modulo/editar/{curso.id}/?step=3")

                curso.estado_publicacion = 'PUBLICADO'
                curso.save()
                messages.success(request, f"Â¡MÃ³dulo de InducciÃ³n '{curso.titulo}' publicado exitosamente!")
                return redirect('onboarding_admin')

    return render(request, 'intranet/lms/crear_curso_stepper.html', {
        'curso': curso, 
        'is_induccion': True,
        'action_url': reverse('editar_curso_induccion_stepper', args=[curso.id]) if curso else reverse('crear_curso_induccion'),
        'lecciones_count': curso.lecciones.count() if curso else 0,
        'tiene_evaluacion': bool(getattr(curso, 'evaluacion', None)) if curso else False,
    })

@login_required(login_url='login')
@solo_directivos
def crear_ruta_induccion_view(request):
    from intranet.models.rrhh_core import Negocio, Area, Cargo, Colaborador
    from intranet.models.lms import CursoInduccion, RutaInduccion, RutaInduccionModulo
    import json

    if request.method == 'POST':
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        rol_objetivo = request.POST.get('rol_objetivo') or None
        area_objetivo_id = request.POST.get('area_objetivo')
        cartera_objetivo_id = request.POST.get('cartera_objetivo')

        ruta = RutaInduccion.objects.create(
            nombre=nombre,
            descripcion=descripcion,
            rol_objetivo=rol_objetivo,
            area_objetivo=Area.objects.filter(id=area_objetivo_id).first() if area_objetivo_id else None,
            cartera_objetivo=Negocio.objects.filter(id=cartera_objetivo_id).first() if cartera_objetivo_id else None
        )

        cursos_seleccionados = request.POST.get('cursos_seleccionados')
        if cursos_seleccionados:
            try:
                curso_ids = json.loads(cursos_seleccionados)
                for index, curso_id in enumerate(curso_ids, start=1):
                    curso = CursoInduccion.objects.filter(id=curso_id).first()
                    if curso:
                        RutaInduccionModulo.objects.create(
                            ruta=ruta,
                            modulo=curso,
                            orden=index
                        )
            except Exception as e:
                pass
        
        messages.success(request, f"Â¡Ruta '{nombre}' creada exitosamente!")
        return redirect('onboarding_admin')

    # Pasar cursos al frontend para el buscador JS
    cursos_qs = CursoInduccion.objects.filter(estado_publicacion='PUBLICADO', tipo='INDUCCION').select_related('categoria_lms')
    cursos_json = []
    for c in cursos_qs:
        cursos_json.append({
            'id': c.id,
            'titulo': c.titulo,
            'categoria': c.categoria_lms.nombre if c.categoria_lms else 'General',
            'publico_general': c.publico_general,
            'rol_permitido': c.rol_permitido or ''
        })

    return render(request, 'intranet/lms/rutas_crear.html', {
        'roles': Colaborador.ROLES,
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'negocios': Negocio.objects.all(),
        'cursos_json': json.dumps(cursos_json),
        'modo_onboarding': True,
    })

@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def cerrar_encuesta(request, pk):
    encuesta = get_object_or_404(Encuesta, id=pk)
    encuesta.activa = False
    encuesta.save()
    messages.success(request, f"La encuesta '{encuesta.titulo}' ha sido cerrada.")
    return redirect('encuestas_admin')

@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def abrir_encuesta(request, pk):
    encuesta = get_object_or_404(Encuesta, id=pk)
    encuesta.activa = True
    encuesta.save()
    messages.success(request, f"La encuesta '{encuesta.titulo}' ha sido reabierta.")
    return redirect('encuestas_admin')

@login_required(login_url='login')
@solo_directivos
@require_http_methods(["POST"])
def eliminar_encuesta(request, pk):
    encuesta = get_object_or_404(Encuesta, id=pk)
    encuesta.delete()
    messages.success(request, f"La encuesta ha sido eliminada permanentemente.")
    return redirect('encuestas_admin')

# ==========================================
# GESTIÃ“N DETALLADA LMS (CURRÃCULUM Y CATEGORÃAS)
# ==========================================
@login_required(login_url='login')
@solo_directivos
def editar_categoria_lms(request, pk):
    categoria = get_object_or_404(CategoriaModuloLMS, id=pk)
    if request.method == 'POST':
        categoria.nombre = request.POST.get('nombre_categoria')
        categoria.descripcion = request.POST.get('descripcion_categoria')
        categoria.icono = request.POST.get('icono_categoria')
        categoria.color = request.POST.get('color_categoria')
        categoria.save()
        messages.success(request, f"CategorÃ­a '{categoria.nombre}' actualizada.")
    return redirect('gestor_lms')

@login_required(login_url='login')
@solo_directivos
def eliminar_categoria_lms(request, pk):
    categoria = get_object_or_404(CategoriaModuloLMS, id=pk)
    nombre = categoria.nombre
    categoria.delete()
    messages.success(request, f"CategorÃ­a '{nombre}' eliminada correctamente.")
    return redirect('gestor_lms')

@login_required(login_url='login')
@solo_directivos
def curso_curriculum(request, pk):
    curso = get_object_or_404(CursoInduccion, id=pk)
    
    if request.method == 'POST':
        if 'crear_leccion' in request.POST:
            video = request.POST.get('url_video')
            archivo_pdf = request.FILES.get('archivo_pdf')
            if not _leccion_tiene_contenido_minimo(video, archivo_pdf):
                messages.error(request, "Cada clase debe incluir al menos un video o un PDF.")
                return redirect('curso_curriculum', pk=curso.id)

            LeccionCurso.objects.create(
                curso=curso,
                titulo=request.POST.get('titulo'),
                descripcion=request.POST.get('descripcion'),
                url_video=video,
                url_presentacion_canva=request.POST.get('url_presentacion_canva') or None,
                url_simulador=request.POST.get('url_simulador') or None,
                paquete_scorm_url=request.POST.get('paquete_scorm_url') or None,
                archivo_pdf=archivo_pdf,
                orden=request.POST.get('orden', 1)
            )
            messages.success(request, f"Â¡Clase agregada correctamente al curso!")
            return redirect('curso_curriculum', pk=curso.id)
            
        elif 'crear_evaluacion' in request.POST:
            if hasattr(curso, 'evaluacion'):
                messages.error(request, "Este curso ya tiene una evaluaciÃ³n.")
            else:
                EvaluacionCurso.objects.create(
                    curso=curso,
                    titulo=request.POST.get('titulo'),
                    instrucciones=request.POST.get('instrucciones', ''),
                    puntaje_maximo=request.POST.get('puntaje_maximo', 20.00),
                    puntaje_aprobatorio=request.POST.get('puntaje_aprobatorio', 14.00),
                    preguntas_a_mostrar=request.POST.get('preguntas_a_mostrar', 10),
                    orden_aleatorio=request.POST.get('orden_aleatorio') == 'on',
                    tiempo_limite_minutos=request.POST.get('tiempo_limite_minutos', 0),
                    puntos_premio=request.POST.get('puntos_premio', 50),
                    intentos_maximos=request.POST.get('intentos_maximos') or 0,
                    mostrar_resultado_inmediato=request.POST.get('mostrar_resultado_inmediato') == 'on',
                    permitir_revision_respuestas=request.POST.get('permitir_revision_respuestas') == 'on',
                    retroalimentacion_final=request.POST.get('retroalimentacion_final', '').strip(),
                )
                messages.success(request, "Â¡Examen creado correctamente!")
            return redirect('curso_curriculum', pk=curso.id)

    return render(request, 'intranet/lms/curso_curriculum.html', {
        'curso': curso,
        'lecciones': curso.lecciones.all().order_by('orden'),
        'categorias_lms': CategoriaModuloLMS.objects.filter(activa=True).order_by('nombre'),
        'negocios': Negocio.objects.all(),
        'areas': Area.objects.filter(activa=True).order_by('nombre'),
        'cargos': Cargo.objects.filter(activa=True).select_related('area').order_by('nombre'),
        'roles': Colaborador.ROLES,
        'cursos_referencia': CursoInduccion.objects.filter(tipo=curso.tipo).exclude(id=curso.id).order_by('titulo'),
    })

@login_required(login_url='login')
@solo_directivos
def editar_leccion_lms(request, pk):
    leccion = get_object_or_404(LeccionCurso, id=pk)
    curso_id = leccion.curso.id
    if request.method == 'POST':
        leccion.titulo = request.POST.get('titulo')
        leccion.descripcion = request.POST.get('descripcion')
        leccion.url_video = request.POST.get('url_video')
        leccion.url_presentacion_canva = request.POST.get('url_presentacion_canva') or None
        leccion.url_simulador = request.POST.get('url_simulador') or None
        leccion.paquete_scorm_url = request.POST.get('paquete_scorm_url') or None
        leccion.orden = request.POST.get('orden', leccion.orden)
        nuevo_pdf = request.FILES.get('archivo_pdf')
        if nuevo_pdf:
            leccion.archivo_pdf = request.FILES.get('archivo_pdf')

        if not _leccion_tiene_contenido_minimo(leccion.url_video, leccion.archivo_pdf):
            messages.error(request, "Cada clase debe incluir al menos un video o un PDF.")
            return redirect('curso_curriculum', pk=curso_id)

        leccion.save()
        messages.success(request, f"LecciÃ³n '{leccion.titulo}' actualizada.")
    return redirect('curso_curriculum', pk=curso_id)

@login_required(login_url='login')
@solo_directivos
def eliminar_leccion_lms(request, pk):
    leccion = get_object_or_404(LeccionCurso, id=pk)
    curso_id = leccion.curso.id
    titulo = leccion.titulo
    leccion.delete()
    messages.success(request, f"LecciÃ³n '{titulo}' eliminada.")
    return redirect('curso_curriculum', pk=curso_id)

@login_required(login_url='login')
@solo_directivos
def eliminar_evaluacion_lms(request, pk):
    evaluacion = get_object_or_404(EvaluacionCurso, id=pk)
    curso_id = evaluacion.curso.id
    evaluacion.delete()
    messages.success(request, "EvaluaciÃ³n eliminada correctamente.")
    return redirect('curso_curriculum', pk=curso_id)

